"""
XUSE 厦大固态实验设备驱动
继承自 OPC UA 通讯基类，实现具体的设备动作函数
"""

import json
import time
import traceback
from typing import Optional
import os
import threading

# 导入日志类
from unilabos.utils.log import logger
import logging

from unilabos.registry.decorators import (
    action,
    device,
    not_action,
    topic_config,
    NodeType,
)

# 导入通讯基类
from unilabos.devices.workstation.XUSE.base_opcua_client import OpcUaClientWithSubscription

# 导入 deck 资源树
from unilabos.devices.workstation.XUSE.decks import XUSE_deck

# 导入常量定义
from unilabos.devices.workstation.XUSE.XUSE_CONSTS import RoboticArmTargetPosition_1, RoboticArmPickPlaceCode_1
from unilabos.devices.workstation.XUSE.XUSE_CONSTS import RoboticArmPickPlaceCode_2, RoboticArmTargetPosition_3
from unilabos.devices.workstation.XUSE.XUSE_CONSTS import RoboticArmPickPlaceCode_3
from unilabos.devices.workstation.XUSE.XUSE_CONSTS import OpenCanActionCode, SieveActionCode, ScrapePowderActionCode
from unilabos.devices.workstation.XUSE.XUSE_CONSTS import SmallCrucibleDischargePosition, LargeCrucibleFeedPosition
from unilabos.devices.workstation.XUSE.XUSE_CONSTS import ARM_LOCK_MAP, ARM_STATUS_NODES

# 定义 XUSE 设备通信类
# 包含三个机械臂，一个罐架区，一个加珠区，一个开罐区，一个刮粉区，一个过筛区，一个加粉区，一个球磨区，一个马弗炉区，一个出料区
@device(
    id="XUSE_station",
    category=["XUSE_station"],
    description="厦门大学固态实验工站（XUSE），包含 3 个机械臂、罐架区、加珠区、开罐区、刮粉区、过筛区、加粉区、球磨区、马弗炉区和出料区",
    display_name="XUSE 厦大固态实验工站",
    icon="XUSE_station.png",
    version="1.0.0",
)
class XUSEDevice(OpcUaClientWithSubscription):
    """
    XUSE 设备类
    继承自 OpcUaClientWithSubscription，实现具体的设备动作函数
    """

    # 动作 -> 机械臂编号 映射（定义见 XUSE_CONSTS.ARM_LOCK_MAP）。
    _ARM_LOCK_MAP = ARM_LOCK_MAP

    def __init__(
        self, 
        url: str, 
        deck: Optional[XUSE_deck] = None,
        csv_path: str = None, 
        username: str = None, 
        password: str = None,
        use_subscription: bool = True,
        cache_timeout: float = 5.0,
        subscription_interval: int = 500,
        *args,
        **kwargs,
    ):
        """
        初始化 XUSE 设备
        
        参数:
            url: OPC UA 服务器地址
            deck: XUSE 资源树配置
            csv_path: 节点配置 CSV 文件路径
            username: OPC UA 用户名
            password: OPC UA 密码
            use_subscription: 是否启用订阅模式
            cache_timeout: 缓存超时时间（秒）
            subscription_interval: 订阅发布间隔（毫秒）
        """
        # 调用父类构造函数
        super().__init__(
            url=url,
            username=username,
            password=password,
            use_subscription=use_subscription,
            cache_timeout=cache_timeout,
            subscription_interval=subscription_interval,
            *args,
            **kwargs
        )

        # 处理 deck 参数
        if deck is None or isinstance(deck.get("data") if isinstance(deck, dict) else deck, dict):
            self.deck = XUSE_deck(setup=True)
        else:
            self.deck = deck.get("data") if isinstance(deck, dict) else deck

        if self.deck is None:
            raise ValueError("Deck 配置不能为空")

        # 统计仓库信息
        if hasattr(self.deck, "children"):
            warehouse_count = len(self.deck.children)
            logger.info(f"Deck 初始化完成，加载 {warehouse_count} 个资源")

        # 机械臂暂存载具（动作间转移物料用）：{arm_id: carrier}
        self._held_carriers = {}

        # 如果提供了 CSV 路径，则直接加载节点
        if csv_path:
            self.load_nodes_from_csv(csv_path)

        # 机械臂状态本地缓存 + 后台轮询线程。
        # 原因：6 个状态方法会被 ROS 各自的定时器周期调用，如果直接走 get_node_value
        # （共享 OPC 锁、可能因动作占用而阻塞），部分发布回调会卡住，导致对应 topic 不发布、
        # host 扫不到、前端状态显示不全。这里用单一后台线程统一刷新缓存，状态方法只读缓存
        # 即时返回（非阻塞、永不 None），保证 6 个状态都能稳定发布。
        self._arm_status_nodes = list(ARM_STATUS_NODES)
        # 启动时先同步读一次真实值初始化缓存（读失败才退回 False 兜底），
        # 保证 6 个状态从一开始就是 OPC UA 的真实状态，且始终是具体 bool（永不 None、永远发布）。
        self._arm_status_cache = {}
        for _n in self._arm_status_nodes:
            try:
                self._arm_status_cache[_n] = bool(self.get_node_value(_n))
            except Exception:
                self._arm_status_cache[_n] = False
        self._arm_status_poller_stop = threading.Event()
        self._arm_status_thread = threading.Thread(
            target=self._arm_status_poll_loop, name="XUSEArmStatusPoller", daemon=True
        )
        self._arm_status_thread.start()

        # 机械臂线程锁：机器人 1/2/3 各一把，保证同一机械臂同一时间只执行一个动作。
        # 仿照 AI4M，用线程锁串行化同一机械臂的动作；用 RLock 允许同线程内嵌套调用（如编排动作）。
        self._arm_locks = {1: threading.RLock(), 2: threading.RLock(), 3: threading.RLock()}
        for _mname, _arm_id in self._ARM_LOCK_MAP.items():
            _orig = getattr(self, _mname, None)
            if callable(_orig):
                setattr(self, _mname, self._make_arm_locked(_orig, _arm_id))
            else:
                logger.warning(f"机械臂线程锁包裹失败，方法不存在: {_mname}")

    @not_action
    def _make_arm_locked(self, func, arm_id: int):
        """把动作方法包裹成"先获取对应机械臂线程锁，执行后释放"的版本（实例级替换）。"""
        import functools

        lock = self._arm_locks[arm_id]

        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            lock.acquire()
            logger.info(f"[机械臂{arm_id}] 已获取线程锁: {func.__name__}")
            try:
                return func(*args, **kwargs)
            finally:
                lock.release()
                logger.info(f"[机械臂{arm_id}] 已释放线程锁: {func.__name__}")

        return wrapper

    @not_action
    def _arm_status_poll_loop(self):
        """后台每秒轮询 6 个机械臂状态节点刷新缓存（容错，读取失败保留上次值）。

        缓存在启动时已用真实值初始化，状态方法始终返回具体 bool、6 个 topic 都能稳定发布。
        """
        while not self._arm_status_poller_stop.is_set():
            for node_name in self._arm_status_nodes:
                try:
                    self._arm_status_cache[node_name] = bool(self.get_node_value(node_name))
                except Exception:
                    pass
            self._arm_status_poller_stop.wait(1.0)

    @not_action
    def post_init(self, ros_node):
        """ROS2 节点就绪后的初始化：本地注册并上传 deck 到云端"""
        if not (hasattr(self, "deck") and self.deck):
            return

        if not (hasattr(ros_node, "resource_tracker") and ros_node.resource_tracker):
            logger.warning("resource_tracker 不存在，无法注册 deck")
            return

        # 保存 ros_node 引用
        self._ros_node = ros_node

        # 1. 本地注册（必需）
        ros_node.resource_tracker.add_resource(self.deck)

        # 2. 上传云端
        try:
            from unilabos.ros.nodes.base_device_node import ROS2DeviceNode
            ROS2DeviceNode.run_async_func(
                ros_node.update_resource,
                True,
                resources=[self.deck]
            )
            logger.info("Deck 已上传到云端")
        except Exception as e:
            logger.error(f"上传失败: {e}")

    # =================== 前端物料转移辅助方法 ===================

    @not_action
    def _sync_deck_to_frontend(self) -> None:
        """将 deck 资源树同步到前端"""
        if hasattr(self, "_ros_node") and self._ros_node:
            try:
                from unilabos.ros.nodes.base_device_node import ROS2DeviceNode
                ROS2DeviceNode.run_async_func(self._ros_node.update_resource, True, resources=[self.deck])
                logger.info("✓ 已同步资源更新到前端")
            except Exception as e:
                logger.warning(f"前端资源更新失败: {e}")

    @not_action
    def _find_carrier_in_warehouse(self, warehouse):
        """在堆栈中查找已有载具：兼容 直接位于 site / 嵌套在 ResourceHolder / children 兜底。

        返回找到的载具（保留其原始名称），未找到返回 None。
        """
        if warehouse is None:
            return None
        # 1) 直接位于 site 上的载具
        for s in (warehouse.sites or []):
            if s is not None and type(s).__name__ != "ResourceHolder":
                return s
        # 2) 嵌套在 ResourceHolder 内部的载具
        for s in (warehouse.sites or []):
            if s is not None and type(s).__name__ == "ResourceHolder":
                inner = getattr(s, "resource", None)
                if inner is not None:
                    return inner
        # 3) 兜底：遍历 children 找载具
        for ch in getattr(warehouse, "children", []) or []:
            if getattr(ch, "category", "") == "bottle_carrier" or "Carrier" in type(ch).__name__:
                return ch
        return None

    @not_action
    def _pick_carrier_from_warehouse(self, warehouse_name: str, arm_id: int):
        """从指定 1x1 堆栈取走载具：解绑 → 暂存到机械臂 → 同步前端。

        - 若堆栈已有物料：取走并保留其原始名称（不会改名）。
        - 若堆栈为空：不做任何转移，也不新建物料（动作照常进行，不报错）。

        参数:
        - warehouse_name: 源堆栈名称（如「开盖区」）
        - arm_id: 机械臂编号（暂存键）
        """
        warehouse = self.deck.warehouses.get(warehouse_name) if getattr(self, "deck", None) else None
        carrier = self._find_carrier_in_warehouse(warehouse)
        if carrier is not None:
            try:
                parent = getattr(carrier, "parent", None) or warehouse
                parent.unassign_child_resource(carrier)
                logger.info(f"✓ 已从「{warehouse_name}」取走载具 {carrier.name}（保留原名）")
            except Exception as e:
                logger.warning(f"从「{warehouse_name}」解绑载具失败（不影响硬件操作）: {e}")
            self._held_carriers[arm_id] = carrier
            self._sync_deck_to_frontend()
        else:
            logger.info(f"「{warehouse_name}」无物料，跳过物料转移")
            self._held_carriers[arm_id] = None
        return carrier

    @not_action
    def _place_carrier_to_warehouse(self, warehouse_name: str, arm_id: int):
        """将机械臂暂存载具放入指定 1x1 堆栈：绑定 → 清空暂存 → 同步前端。

        - 若机械臂有暂存物料：放入目标堆栈。
        - 若没有暂存物料：不做任何转移，也不新建物料（动作照常进行，不报错）。

        参数:
        - warehouse_name: 目标堆栈名称（如「加样区」）
        - arm_id: 机械臂编号（暂存键）
        """
        warehouse = self.deck.warehouses.get(warehouse_name) if getattr(self, "deck", None) else None
        carrier = self._held_carriers.get(arm_id)
        if warehouse is not None and carrier is not None:
            try:
                site_idx = 0
                site_key = list(warehouse._ordering.keys())[site_idx]
                location = warehouse.child_locations[site_key]
                warehouse.assign_child_resource(carrier, location=location, spot=site_idx)
                logger.info(f"✓ 已将载具 {carrier.name} 放入「{warehouse_name}」")
            except Exception as e:
                logger.warning(f"将载具放入「{warehouse_name}」失败（不影响硬件操作）: {e}")
            self._held_carriers[arm_id] = None
            self._sync_deck_to_frontend()
        else:
            logger.info(f"机械臂{arm_id}无暂存物料，跳过物料转移")
        return carrier

    @not_action
    def _pick_carrier_from_warehouse_at(self, warehouse_name: str, site_key, arm_id: int):
        """从多位堆栈指定位 site_key 取走载具，暂存到机械臂（无物料则跳过，不报错）。

        参数:
        - warehouse_name: 源堆栈名称（如「球磨罐仓库」）
        - site_key: 堆栈内位键（如「1-1」「C-1」「2」）
        - arm_id: 机械臂编号（暂存键）
        """
        warehouse = self.deck.warehouses.get(warehouse_name) if getattr(self, "deck", None) else None
        carrier = None
        if warehouse is not None:
            try:
                site_idx = list(warehouse._ordering.keys()).index(str(site_key))
                site = warehouse.sites[site_idx]
                if site is not None and type(site).__name__ == "ResourceHolder":
                    holder = site
                    carrier = getattr(site, "resource", None)
                else:
                    holder = None
                    carrier = site  # 直接位于位上的载具
                if carrier is not None:
                    parent = getattr(carrier, "parent", None) or holder or warehouse
                    parent.unassign_child_resource(carrier)
                    logger.info(f"✓ 已从「{warehouse_name}」[{site_key}] 取走载具 {carrier.name}（保留原名）")
            except Exception as e:
                logger.warning(f"从「{warehouse_name}」[{site_key}] 取载具失败（不影响硬件操作）: {e}")
                carrier = None
        if carrier is not None:
            self._held_carriers[arm_id] = carrier
            self._sync_deck_to_frontend()
        else:
            logger.info(f"「{warehouse_name}」[{site_key}] 无物料，跳过物料转移")
            self._held_carriers[arm_id] = None
        return carrier

    @not_action
    def _place_carrier_to_warehouse_at(self, warehouse_name: str, site_key, arm_id: int):
        """将机械臂暂存载具放入多位堆栈指定位 site_key（无暂存则跳过，不报错）。

        参数:
        - warehouse_name: 目标堆栈名称（如「过筛区」）
        - site_key: 堆栈内位键（如「1」「1-1」「D-1」）
        - arm_id: 机械臂编号（暂存键）
        """
        warehouse = self.deck.warehouses.get(warehouse_name) if getattr(self, "deck", None) else None
        carrier = self._held_carriers.get(arm_id)
        if warehouse is not None and carrier is not None:
            try:
                site_idx = list(warehouse._ordering.keys()).index(str(site_key))
                location = warehouse.child_locations[str(site_key)]
                warehouse.assign_child_resource(carrier, location=location, spot=site_idx)
                logger.info(f"✓ 已将载具 {carrier.name} 放入「{warehouse_name}」[{site_key}]")
            except Exception as e:
                logger.warning(f"将载具放入「{warehouse_name}」[{site_key}] 失败（不影响硬件操作）: {e}")
            self._held_carriers[arm_id] = None
            self._sync_deck_to_frontend()
        else:
            logger.info(f"机械臂{arm_id}无暂存物料，跳过物料转移")
        return carrier

    @not_action
    def _can_rack_site_key(self, position: int) -> str:
        """罐架位置号(1-32) → 球磨罐仓库位键（按行：1→1-1, 9→2-1, 32→4-8）。"""
        return f"{(position - 1) // 8 + 1}-{(position - 1) % 8 + 1}"

    @not_action
    def _small_crucible_rack_site_key(self, position: int) -> str:
        """坩埚位置号(1-20) → 小坩埚仓库位键（1→1-1, 10→1-10, 11→2-1, 20→2-10）。"""
        return f"{(position - 1) // 10 + 1}-{(position - 1) % 10 + 1}"

    @not_action
    def _wait_condition(self, predicate, timeout: float = 3.0, interval: float = 0.1) -> bool:
        """持续检测占位条件：满足返回 True；超过 timeout（默认 3 秒）仍不满足返回 False。

        用于所有占位检测：条件不满足时不立即报错，而是持续轮询到超时再交由调用方处理。
        """
        deadline = time.time() + timeout
        while True:
            try:
                if predicate():
                    return True
            except Exception as e:
                logger.warning(f"占位检测异常，重试中: {e}")
            if time.time() >= deadline:
                return False
            time.sleep(interval)

    # 初始化工站
    @action(
        always_free=True,
        node_type=NodeType.MANUAL_CONFIRM,
        placeholder_keys={"assignee_user_ids": "unilabos_manual_confirm"},
        goal_default={"timeout_seconds": 3600, "assignee_user_ids": []},
        feedback_interval=300,
        description="工站初始化（人工确认节点：确认通过后停止机械臂触发，触发工站初始化，等待初始化完成）",
    )
    def trigger_init(
        self,
        timeout_seconds: int = 3600,
        assignee_user_ids: Optional[list] = None,
        **kwargs,
    ) -> dict:
        """
        初始化函数（人工确认节点：云端确认通过后才会执行）：
        - 停止 3 个机械臂触发
        - 触发工站初始化
        - 等待初始化完成

        Args:
            timeout_seconds[超时时间]: 人工确认超时时间，单位秒。
            assignee_user_ids[确认人]: 指定处理人工确认任务的用户 ID 列表。

        Returns:
            dict: 包含 success 和 message
        """
        logger.info("停止机械臂触发...")
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)
        self.set_node_value("Robotic_Arm_Action_Trigger_3", False)

        logger.info("进行初始化...")
        self.set_node_value("Station_Initialize_Complete", False)
        time.sleep(1.0) 
        self.set_node_value("Station_Initialize", True) 
        time.sleep(1.0)
        if self._wait_until_true("Station_Initialize_Complete", description="初始化工站"):
            logger.info("初始化工站成功")
            self.set_node_value("Station_Initialize", False)
            return {
                "success": True,
                "message": "初始化工站成功",
            }
        else:
            logger.error("初始化工站失败")
            self.set_node_value("Station_Initialize", False)
            raise ValueError("初始化工站失败")

    # 加样单元初始化
    @action(description="加样单元初始化（触发 Add_Sample_Initialize，等待 Add_Sample_Initialize_Complete）")
    def trigger_add_powder_init(self) -> dict:
        """
        加样单元初始化：
        - 先复位完成标志 Add_Sample_Initialize_Complete
        - 触发 Add_Sample_Initialize（上升沿）
        - 等待 Add_Sample_Initialize_Complete 变为 True
        - 复位 Add_Sample_Initialize，返回成功

        本动作只负责加样单元的自初始化，不影响机械臂等其它模块；与 trigger_init（整站初始化）
        独立。如需在整站初始化前后单独重置加样单元，可调本动作。

        Returns:
            dict: 包含 success 和 message
        """
        logger.info("开始加样单元初始化...")
        # 先复位完成标志与触发标志，避免残留信号
        self.set_node_value("Add_Sample_Initialize_Complete", False)
        time.sleep(1.0)
        self.set_node_value("Add_Sample_Initialize", True)  # 上升沿触发
        time.sleep(1.0)
        if self._wait_until_true("Add_Sample_Initialize_Complete", description="加样初始化"):
            logger.info("加样单元初始化成功")
            self.set_node_value("Add_Sample_Initialize", False)  # 复位触发
            return {
                "success": True,
                "message": "加样单元初始化成功",
            }
        else:
            logger.error("加样单元初始化失败")
            self.set_node_value("Add_Sample_Initialize", False)
            raise ValueError("加样单元初始化失败")

    def is_robotic_arm_idle(self, arm_id: int) -> bool:
        """
        检查机械臂是否空闲
        
        参数:
            arm_id: 机械臂ID,1,2,3
        
        Returns:
            bool: 如果机械臂空闲，返回True，否则返回False
        """
        if arm_id not in [1, 2, 3]:
            raise ValueError("机械臂ID必须为1,2,3")
        return self.get_node_value(f"Robotic_Arm_Idle_{arm_id}")

    # =================== 设备节点状态（前端显示：机械臂空闲/故障） ===================

    @not_action
    def _read_bool_node(self, node_name: str) -> bool:
        """读取机械臂状态缓存：非阻塞、永不抛错、永不返回 None（默认 False）。"""
        return bool(self._arm_status_cache.get(node_name, False))

    @topic_config(period=1.0)
    def robotic_arm_1_idle(self) -> bool:
        """机械臂1空闲状态"""
        return self._read_bool_node("Robotic_Arm_Idle_1")

    @topic_config(period=1.0)
    def robotic_arm_2_idle(self) -> bool:
        """机械臂2空闲状态"""
        return self._read_bool_node("Robotic_Arm_Idle_2")

    @topic_config(period=1.0)
    def robotic_arm_3_idle(self) -> bool:
        """机械臂3空闲状态"""
        return self._read_bool_node("Robotic_Arm_Idle_3")

    @topic_config(period=1.0)
    def robotic_arm_1_fault(self) -> bool:
        """机械臂1故障状态"""
        return self._read_bool_node("Robotic_Arm_Fault_1")

    @topic_config(period=1.0)
    def robotic_arm_2_fault(self) -> bool:
        """机械臂2故障状态"""
        return self._read_bool_node("Robotic_Arm_Fault_2")

    @topic_config(period=1.0)
    def robotic_arm_3_fault(self) -> bool:
        """机械臂3故障状态"""
        return self._read_bool_node("Robotic_Arm_Fault_3")

    @not_action
    def is_open_can_upper_lid_occupied(self) -> bool:
        """
        检查开罐上盖是否占位
        
        Returns:
            bool: 如果开罐上盖占位，返回True，否则返回False
        """
        return self.get_node_value("Open_Can_Upper_Lid_Occupied")
    
    @not_action
    def is_open_can_body_occupied(self) -> bool:
        """
        检查开罐主体是否占位
        
        Returns:
            bool: 如果开罐主体占位，返回True，否则返回False
        """
        return self.get_node_value("Open_Can_Body_Occupied")
    
    @not_action
    def is_add_sample_occupied(self) -> bool:
        """
        检查加样是否占位
        
        Returns:
            bool: 如果加样占位，返回True，否则返回False
        """
        return self.get_node_value("Add_Sample_Occupied")
    
    @not_action
    def is_add_bead_occupied(self) -> bool:
        """
        检查加珠是否占位
        
        Returns:
            bool: 如果加珠占位，返回True，否则返回False
        """
        return self.get_node_value("Add_Bead_Occupied")
    
    @not_action
    def is_ball_mill_occupied(self, mill_position: int) -> bool:
        """
        检查球磨区是否占位
        
        参数:
            mill_position: 球磨区位置
        
        Returns:
            bool: 如果球磨区占位，返回True，否则返回False
        """
        return self.get_node_value(f"Ball_Mill_Occupied_{mill_position}")
    
    @not_action
    def is_sieve_can_occupied(self) -> bool:
        """
        检查过筛区球磨罐是否占位
        
        Returns:
            bool: 如果过筛区球磨罐占位，返回True，否则返回False
        """
        return self.get_node_value("Sieve_Can_Occupied")
    
    @not_action
    def is_sieve_crucible_occupied(self) -> bool:
        """
        检查过筛区小坩埚是否占位
        
        Returns:
            bool: 如果过筛区小坩埚占位，返回True，否则返回False
        """
        return self.get_node_value("Sieve_Crucible_Occupied")
    
    @not_action
    def is_sieve_funnel_occupied(self) -> bool:
        """
        检查过筛区漏斗是否占位
        
        Returns:
            bool: 如果过筛区漏斗占位，返回True，否则返回False
        """
        return self.get_node_value("Sieve_Funnel_Occupied")
    
    @not_action
    def is_scrape_occupied(self) -> bool:
        """
        检查刮粉区是否占位
        
        Returns:
            bool: 如果刮粉区占位，返回True，否则返回False
        """
        return self.get_node_value("Scrape_Powder_Occupied")

    @not_action
    def is_can_rack_occupied(self, position: int) -> bool:
        """
        检查罐架区（球磨罐仓库）指定位置是否占位（对应 ROBOT_1_occupy[position]）。

        参数:
            position: 罐架位置，1-32

        Returns:
            bool: 占位返回 True，否则返回 False
        """
        return bool(self.get_node_value(f"Can_Rack_Occupied_{position}"))

    @not_action
    def is_crucible_rack_occupied(self, code: int) -> bool:
        """
        检查坩埚架区指定取放代码位置是否占位（对应 ROBOT_2_occupy[code]）。

        说明：code 为机械臂2取放代码，放漏斗为 21-28、取漏斗为 31-38。

        参数:
            code: 机械臂2取放代码

        Returns:
            bool: 占位返回 True，否则返回 False
        """
        return self.get_node_value(f"Crucible_Rack_Occupied_{code}")
    
    @not_action
    def is_small_crucible_discharge_occupied(self, position: int) -> bool:
        """
        检查小坩埚出料位指定位置是否占位
        
        参数:
            position: 小坩埚出料位编号，1-4
        
        Returns:
            bool: 如果该位置占位，返回True，否则返回False
        """
        if position not in [1, 2, 3, 4]:
            raise ValueError(f"小坩埚出料位编号必须在 1-4 范围内，当前值: {position}")
        return self.get_node_value(f"Small_Crucible_Discharge_Occupied_{position}")
    
    @not_action
    def get_small_crucible_discharge_current_position(self) -> int:
        """
        获取小坩埚出料当前位置
        
        Returns:
            int: 小坩埚出料当前位置
        """
        return self.get_node_value("Small_Crucible_Discharge_Current_Position")
    
    @not_action
    def get_large_crucible_feed_current_position(self) -> int:
        """
        获取大坩埚入料当前位置
        
        Returns:
            int: 大坩埚入料当前位置
        """
        return self.get_node_value("Large_Crucible_Feed_Current_Position")
    
    @not_action
    def is_muffle_furnace_occupied(self, muffle_furnace_position: int) -> bool:
        """
        检查马弗炉是否占位
        
        参数:
            muffle_furnace_position: 马弗炉位置
        
        Returns:
            bool: 如果马弗炉占位，返回True，否则返回False
        """
        return self.get_node_value(f"Muffle_Furnace_Occupied_{muffle_furnace_position}")
    
    @not_action
    def is_upper_product_rack_occupied(self) -> bool:
        """
        检查上成品架是否占位
        
        Returns:
            bool: 如果上成品架占位，返回True，否则返回False
        """
        return self.get_node_value("Upper_Product_Rack_Occupied")
    
    @not_action
    def is_lower_product_rack_occupied(self) -> bool:
        """
        检查下成品架是否占位
        
        Returns:
            bool: 如果下成品架占位，返回True，否则返回False
        """
        return self.get_node_value("Lower_Product_Rack_Occupied")
    
    @action()
    def pick_can_from_can_rack(self, rack_position: int) -> dict:
        """
        从罐架区取球磨罐
        - 检查机械臂1是否空闲
        - 设置从罐架位置rack_position处抓取球磨罐
        - 等待从罐架抓取球磨罐完成
        - 返回成功
        
        参数:
            rack_position: 罐架区位置
        
        Returns:
            dict: 包含 success 和 message
        """
        logger.info(f"从罐架区取球磨罐，位置：{rack_position}")
        MIN_RACK_POSITION = 1
        MAX_RACK_POSITION = RoboticArmPickPlaceCode_1.PICK_CAN_RACK_END - RoboticArmPickPlaceCode_1.PICK_CAN_RACK_START + 1
        if rack_position < MIN_RACK_POSITION or rack_position > MAX_RACK_POSITION:
            error_msg = "罐架位置错误"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")

        if not self._wait_condition(lambda: self.is_can_rack_occupied(rack_position)):
            error_msg = f"罐架位置{rack_position}无球磨罐，无法抓取"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.CAN_RACK_POSITION) # 设置机械臂目标位置为罐架
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.PICK_CAN_RACK_START + rack_position - 1) # 设置罐架位置
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="从罐架抓取球磨罐完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="从罐架抓取球磨罐完成"): # 等待完成状态复位
                logger.info("从罐架区取球磨罐完成")
                # 前端物料转移：从「球磨罐仓库」对应位取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse_at("球磨罐仓库", self._can_rack_site_key(rack_position), arm_id=1)
                return {
                    "success": True,
                    "message": f"从罐架区位置{rack_position}取球磨罐完成",
                }
            else:
                error_msg = "从罐架区取球磨罐失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "从罐架区取球磨罐失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    @action()
    def place_empty_can_to_open_can_position(self) -> dict:
        """
        将空罐放置到开盖区
        - 检查机械臂1是否空闲
        - 检查开罐是否占位
        - 设置将开盖罐磨罐放置到开盖区
        - 等待将开盖罐磨罐放置到开盖区完成
        - 返回成功
        """
        logger.info("将空球磨罐放置到开盖区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")

        if not self._wait_condition(lambda: not (self.is_open_can_upper_lid_occupied() or self.is_open_can_body_occupied())):
            error_msg = "开罐上盖或主体占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开盖区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.OPEN_CAN_NO_POWDER_PLACE_EMPTY_CAN) # 设置开盖区放空罐
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="将球磨罐放置到开盖区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="将球磨罐放置到开盖区完成"): # 等待完成状态复位
                logger.info("将球磨罐放置到开盖区完成")
                # 前端物料转移：将机械臂1暂存载具放入「开盖区」（无暂存则跳过）
                self._place_carrier_to_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": "将球磨罐放置到开盖区完成",
                }
            else:
                error_msg = "将球磨罐放置到开盖区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "将球磨罐放置到开盖区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    @action()
    def open_can_lid(self) -> dict:
        """
        打开罐上盖
        - 检查罐体占位
        - 检查盖子非占位
        - 设置打开开罐上盖
        - 等待请求执行
        - 等待打开开罐上盖完成
        - 返回成功
        """
        logger.info("打开罐上盖...")
        if not self._wait_condition(lambda: self.is_open_can_body_occupied()):
            error_msg = "开罐主体未占位，无法打开"
            logger.error(error_msg)
            raise ValueError(error_msg)

        if not self._wait_condition(lambda: not (self.is_open_can_upper_lid_occupied())):
            error_msg = "开罐上盖已占位，无法打开"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if self._wait_until_true("Open_Can_Request_Process", description="打开上盖请求"):
            logger.info("接收到打开上盖请求")
            self.set_node_value("Open_Can_Action_Control_Code", OpenCanActionCode.OPEN_CAN_LID) # 设置打开开罐上盖
            self.set_node_value("Open_Can_Start_Process", True) # 开始加工
            if self._wait_until_true("Open_Can_Process_Complete", description="打开上盖完成"):
                logger.info("打开上盖完成")
                self.set_node_value("Open_Can_Action_Control_Code", OpenCanActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Open_Can_Start_Process", False) # 复位加工
                return {
                    "success": True,
                    "message": "打开上盖完成",
                }
            else:
                logger.error("打开上盖失败，动作超时")
                self.set_node_value("Open_Can_Action_Control_Code", OpenCanActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Open_Can_Start_Process", False) # 复位加工
                raise ValueError("打开上盖失败，动作超时")
        else:
            error_msg = "打开上盖失败，未收到开盖请求"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        
    @action()
    def pick_empty_can_from_open_can_position(self) -> dict:
        """
        从开盖区抓取空罐
        - 检查机械臂1是否空闲
        - 检查开罐是否占位
        - 设置从开盖区抓取球磨罐
        - 等待从开盖区抓取空罐完成
        - 返回成功
        """
        logger.info("从开盖区抓取空罐...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if not self._wait_condition(lambda: self.is_open_can_body_occupied()):
            error_msg = "开罐主体未占位，无法抓取"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开盖区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.OPEN_CAN_NO_POWDER_PICK_BASE) # 设置开盖区取底座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="从开盖区抓取球磨罐完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="从开盖区抓取球磨罐完成"): # 等待完成状态复位
                logger.info("从开盖区抓取球磨罐完成")
                # 前端物料转移：从「开盖区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": "从开盖区抓取球磨罐完成",
                }
            else:
                error_msg = "从开盖区抓取球磨罐失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "从开盖区抓取球磨罐失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)

    @action()
    def place_can_to_add_powder_position(self) -> dict:
        """
        将罐体放置到加粉区
        - 检查机械臂1是否空闲
        - 检查加样是否占位
        - 等待放置到加粉区完成
        - 返回成功
        """
        logger.info("将罐体放置到加粉区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if not self._wait_condition(lambda: not (self.is_add_sample_occupied())):
            error_msg = "加样占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.ADD_POWDER_POSITION) # 设置机械臂目标位置为加粉区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.ADD_POWDER_PLACE_BASE) # 设置加粉区放底座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="将罐体放置到加粉区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="将罐体放置到加粉区完成"): # 等待完成状态复位
                logger.info("将罐体放置到加粉区完成")
                # 前端物料转移：将机械臂1暂存载具放入「加样区」（无暂存则跳过）
                self._place_carrier_to_warehouse("加样区", arm_id=1)
                return {
                    "success": True,
                    "message": "将罐体放置到加粉区完成",
                }
            else:
                error_msg = "将罐体放置到加粉区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "将罐体放置到加粉区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    @action()
    def add_powder(self, check_can_occupied: bool = True) -> dict:
        """
        加样（加粉）—— 只触发加粉动作，不含参数下发。

        本动作只负责触发一次加粉：
        - （可选）等待罐体占位
        - 等待 Add_Sample_Request_Process 上升沿
        - 触发 Add_Sample_Start_Process
        - 等待 Add_Sample_Process_Complete 后复位

        所有加样相关参数（单值参数：粉末名称/位置号/重量/震荡最高速度/粉末号；
        数组参数：1ML/500NL 的开口量/落粉均速/旋转速度/提前停止量）
        请通过 set_add_powder_params 加载 xlsx 在本动作之前下发。

        Args:
            check_can_occupied[是否检查罐体占位]: True=加粉前等待并校验罐体占位；False=跳过占位检查。
        """
        logger.info(f"加粉...（check_can_occupied={check_can_occupied}）")

        if check_can_occupied:
            if not self._wait_condition(lambda: self.is_add_sample_occupied()):
                error_msg = "没有罐体，无法加粉"
                logger.error(error_msg)
                raise ValueError(error_msg)
            logger.info("有罐体，开始加粉...")
        else:
            logger.info("跳过罐体占位检查，直接开始加粉...")

        if self._wait_until_true("Add_Sample_Request_Process", description="加样请求加工"):
            logger.info("接收到加样请求加工")
            self.set_node_value("Add_Sample_Start_Process", True)  # 开始加工
            if self._wait_until_true("Add_Sample_Process_Complete", description="加样加工完成"):
                logger.info("加样加工完成")
                self.set_node_value("Add_Sample_Start_Process", False)  # 复位加工
                return {
                    "success": True,
                    "message": "加样加工完成",
                }
            else:
                logger.error("加样加工失败，动作超时")
                self.set_node_value("Add_Sample_Start_Process", False)  # 复位加工
                raise ValueError("加样加工失败，动作超时")
        else:
            error_msg = "加样失败，未收到加样请求"
            logger.error(error_msg)
            raise ValueError(error_msg)

    
    @action()
    def pick_can_from_add_powder_position(self) -> dict:
        """
        从加粉区取罐体
        - 检查机械臂1是否空闲
        - 检查加样是否占位
        - 等待取罐体完成
        - 返回成功
        """
        logger.info("从加粉区取罐体...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if not self._wait_condition(lambda: self.is_add_sample_occupied()):
            error_msg = "加样未占位，无法取罐体"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.ADD_POWDER_POSITION) # 设置机械臂目标位置为加粉区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.ADD_POWDER_PICK_BASE) # 设置加粉区取底座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="从加粉区取罐体完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="从加粉区取罐体完成"): # 等待完成状态复位
                logger.info("从加粉区取罐体完成")
                # 前端物料转移：从「加样区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("加样区", arm_id=1)
                return {
                    "success": True,
                    "message": "从加粉区取罐体完成",
                }
            else:
                error_msg = "从加粉区取罐体失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "从加粉区取罐体失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_can_to_add_bead_position(self) -> dict:
        """
        将罐体放置到加珠区
        - 检查机械臂1是否空闲
        - 检查加珠是否占位
        - 等待放置到加珠区完成
        - 返回成功
        """
        logger.info("将罐体放置到加珠区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if not self._wait_condition(lambda: not (self.is_add_bead_occupied())):
            error_msg = "加珠未占位，无法放置罐体"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.ADD_BEAD_POSITION) # 设置机械臂目标位置为加珠区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.ADD_BEAD_PLACE_BASE) # 设置加珠区放底座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="将罐体放置到加珠区成功"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="将罐体放置到加珠区成功"): # 等待完成状态复位
                logger.info("将罐体放置到加珠区成功")
                # 前端物料转移：将机械臂1暂存载具放入「加珠区」（无暂存则跳过）
                self._place_carrier_to_warehouse("加珠区", arm_id=1)
                return {
                    "success": True,
                    "message": "将罐体放置到加珠区成功",
                }
            else:
                error_msg = "将罐体放置到加珠区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "将罐体放置到加珠区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def add_bead(self) -> dict:
        """
        进行加珠操作
        - 检查加珠是否占位
        - 等待加珠完成
        - 返回成功
        """
        logger.info("加珠...")
        if not self._wait_condition(lambda: self.is_add_bead_occupied()):
            error_msg = "加珠未占位，无法加珠"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if self._wait_until_true("Add_Bead_Request_Process", description="加珠请求加工"):
            logger.info("接收到加珠请求加工")
            self.set_node_value("Add_Bead_Start_Process", True) # 设置加珠开始
            if self._wait_until_true("Add_Bead_Process_Complete", description="等待加珠完成"):
                logger.info("加珠完成")
                self.set_node_value("Add_Bead_Start_Process", False) # 复位加珠开始
                return {
                    "success": True,
                    "message": "加珠完成",
                }
            else:
                logger.error("加珠失败")
                self.set_node_value("Add_Bead_Start_Process", False) # 复位加珠开始
                raise ValueError("加珠失败，完成复位超时")
        else:
            error_msg = "加珠失败，未收到加珠请求"
            logger.error(error_msg)
            raise ValueError(error_msg)

    
    @action()
    def pick_can_from_add_bead_position(self) -> dict:
        """
        从加珠区取罐体
        - 检查机械臂1是否空闲
        - 检查加珠是否占位
        - 等待取罐体完成
        - 返回成功
        """
        logger.info("从加珠区取罐体...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if not self._wait_condition(lambda: self.is_add_bead_occupied()):
            error_msg = "加珠未占位，无法取罐体"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.ADD_BEAD_POSITION) # 设置机械臂目标位置为加珠区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.ADD_BEAD_PICK_BASE) # 设置加珠区取底座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="从加珠区取罐体成功"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="从加珠区取罐体成功"): # 等待完成状态复位
                logger.info("从加珠区取罐体成功")
                # 前端物料转移：从「加珠区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("加珠区", arm_id=1)
                return {
                    "success": True,
                    "message": "从加珠区取罐体成功",
                }
            else:
                error_msg = "从加珠区取罐体失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "从加珠区取罐体失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
    

    @action()
    def place_can_with_powder_and_bead_to_open_can_position(self) -> dict:
        """
        将带有粉珠的球磨罐放置到开盖区
        - 检查机械臂1是否空闲
        - 检查开罐是否占位
        - 设置将开盖罐磨罐放置到开盖区
        - 等待将开盖罐磨罐放置到开盖区完成
        - 返回成功
        """
        logger.info("将带有粉珠的球磨罐放置到开盖区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")

        if not self._wait_condition(lambda: not (self.is_open_can_body_occupied())):
            error_msg = "开罐主体占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开盖区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.OPEN_CAN_WITH_POWDER_PLACE_BASE) # 设置开盖区放底座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="将球磨罐放置到开盖区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="将球磨罐放置到开盖区完成"): # 等待完成状态复位
                logger.info("将球磨罐放置到开盖区完成")
                # 前端物料转移：将机械臂1暂存载具放入「开盖区」（无暂存则跳过）
                self._place_carrier_to_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": "将球磨罐放置到开盖区完成",
                }
            else:
                error_msg = "将球磨罐放置到开盖区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "将球磨罐放置到开盖区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def close_can_lid(self) -> dict:
        """
        关闭罐上盖
        - 检查罐体占位
        - 检查盖子占位
        - 设置关闭罐上盖
        - 等待请求执行
        - 等待关闭罐上盖完成
        - 返回成功
        """
        logger.info("关闭罐上盖...")
        if not self._wait_condition(lambda: self.is_open_can_body_occupied()):
            error_msg = "开罐主体未占位，无法关盖"
            logger.error(error_msg)
            raise ValueError(error_msg)

        if not self._wait_condition(lambda: self.is_open_can_upper_lid_occupied()):
            error_msg = "开罐上盖未占位，无法关盖"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if self._wait_until_true("Open_Can_Request_Process", description="关闭上盖请求"):
            logger.info("接收到关闭上盖请求")
            self.set_node_value("Open_Can_Action_Control_Code", OpenCanActionCode.CLOSE_CAN_LID) # 设置关闭罐上盖
            self.set_node_value("Open_Can_Start_Process", True) # 开始加工
            if self._wait_until_true("Open_Can_Process_Complete", description="关闭上盖完成"):
                logger.info("关闭上盖完成")
                self.set_node_value("Open_Can_Action_Control_Code", OpenCanActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Open_Can_Start_Process", False) # 复位加工
                return {
                    "success": True,
                    "message": "关闭上盖完成",
                }
            else:
                logger.error("关闭上盖失败")
                self.set_node_value("Open_Can_Action_Control_Code", OpenCanActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Open_Can_Start_Process", False) # 复位加工
                raise ValueError("关闭上盖失败，完成复位超时")
        else:
            error_msg = "关闭上盖失败，未收到关盖请求"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def pick_can_with_powder_and_bead_from_open_can_position(self) -> dict:
        """
        从开盖区抓取带有粉珠的球磨罐
        - 检查机械臂1是否空闲
        - 检查开罐是否占位
        - 设置从开盖区抓取球磨罐
        - 等待从开盖区抓取空罐完成
        - 返回成功
        """
        logger.info("从开盖区抓取带有粉珠的球磨罐...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if not self._wait_condition(lambda: self.is_open_can_body_occupied()):
            error_msg = "开罐主体未占位，无法抓取"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开盖区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.OPEN_CAN_WITH_POWDER_PICK_FULL_CAN) # 设置开盖区取满罐
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description="从开盖区抓取球磨罐完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description="从开盖区抓取球磨罐完成"): # 等待完成状态复位
                logger.info("从开盖区抓取球磨罐完成")
                # 前端物料转移：从「开盖区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": "从开盖区抓取球磨罐完成",
                }
            else:
                error_msg = "从开盖区抓取球磨罐失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "从开盖区抓取球磨罐失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def place_can_to_ball_mill(self, mill_position: int) -> dict:
        """
        将罐体放置到球磨区
        - 检查机械臂1是否空闲
        - 检查球磨区是否占位
        - 设置将开盖罐磨罐放置到球磨区
        - 等待放置罐体完成
        - 返回成功
        """
        logger.info(f"将开盖罐磨罐放置到球磨区{mill_position}...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"球磨区位置{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: not (self.is_ball_mill_occupied(mill_position))):
            error_msg = f"球磨区{mill_position}占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        mill_position_code = RoboticArmPickPlaceCode_1.BALL_MILL_PLACE_CAN_1 + (mill_position - 1)
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.BALL_MILL_POSITION) # 设置机械臂目标位置为球磨区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", mill_position_code) # 设置球磨区放罐
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"向球磨区{mill_position}放罐完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"向球磨区{mill_position}放罐完成"): # 等待完成状态复位
                logger.info(f"向球磨区{mill_position}放罐完成")
                # 前端物料转移：将机械臂1暂存载具放入「球磨区」对应位（无暂存则跳过）
                self._place_carrier_to_warehouse_at("球磨区", str(mill_position), arm_id=1)
                return {
                    "success": True,
                    "message": f"向球磨区{mill_position}放罐完成",
                }
            else:
                error_msg = f"向球磨区{mill_position}放罐失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"向球磨区{mill_position}放罐失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)

    @action()
    def ball_mill(self, require_full: bool = True) -> dict:
        """
        进行球磨
        - 检测球磨区位置是否有球磨罐
        - 启动球磨
        - 等待球磨完成
        - 返回成功

        Args:
            require_full[是否需满4罐]: 是否要求球磨区 4 个位置全部放满才开始加工。
                True=必须满 4 个；False=只要有球磨罐即可开始（默认 True）。
        """
        if require_full:
            for mill_position in [1, 2, 3, 4]:
                if not self._wait_condition(lambda mp=mill_position: self.is_ball_mill_occupied(mp)):
                    error_msg = f"球磨区位置{mill_position}为空，需满 4 个球磨罐才开始加工"
                    logger.error(error_msg)
                    raise ValueError(error_msg)
        else:
            occupied = [mp for mp in [1, 2, 3, 4]
                        if self._wait_condition(lambda m=mp: self.is_ball_mill_occupied(m))]
            if not occupied:
                error_msg = "球磨区无球磨罐，无法球磨"
                logger.error(error_msg)
                raise ValueError(error_msg)
            logger.info(f"球磨区已占位 {occupied}（不要求满 4 个），开始加工")

        if self._wait_until_true("Ball_Mill_Request_Process", description="球磨请求加工"):
            logger.info("收到球磨请求加工")
            self.set_node_value("Ball_Mill_Start_Process", True) # 设置球磨开始
            if self._wait_until_true("Ball_Mill_Process_Complete", description="等待球磨完成"):
                logger.info("球磨完成")
                self.set_node_value("Ball_Mill_Start_Process", False) # 复位球磨开始
                return {
                    "success": True,
                    "message": "球磨完成",
                }
            else:
                self.set_node_value("Ball_Mill_Start_Process", False) # 复位球磨开始
                error_msg = "球磨失败，操作超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = "球磨加工失败，未收到加工请求"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    
    @action()
    def pick_can_from_ball_mill(self, mill_position: int) -> dict:
        """
        从球磨区抓取罐体
        - 检查机械臂1是否空闲
        - 检查球磨区是否为空
        - 设置从球磨区抓取罐体
        - 等待抓取罐体完成
        - 返回成功
        """
        logger.info(f"从球磨区{mill_position}抓取罐体...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"球磨区位置{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: self.is_ball_mill_occupied(mill_position)):
            error_msg = f"球磨区位置{mill_position}为空，无法抓取"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        mill_position_code = RoboticArmPickPlaceCode_1.BALL_MILL_PICK_CAN_1 + (mill_position - 1)
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.BALL_MILL_POSITION) # 设置机械臂目标位置为球磨区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", mill_position_code) # 设置球磨区抓取罐 
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"从球磨区位置{mill_position}抓取罐完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"从球磨区位置{mill_position}抓取罐完成"): # 等待完成状态复位
                logger.info(f"从球磨区位置{mill_position}抓取罐完成")
                # 前端物料转移：从「球磨区」对应位取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse_at("球磨区", str(mill_position), arm_id=1)
                return {
                    "success": True,
                    "message": f"从球磨区位置{mill_position}抓取罐完成",
                }
            else:
                error_msg = f"从球磨区位置{mill_position}抓取罐失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"从球磨区位置{mill_position}抓取罐失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)

    
    @action()
    def place_milled_can_to_open_can_position(self, mill_position: int) -> dict:
        """
        将研磨后球磨罐放到开盖区
        - 检查机械臂1是否空闲
        - 检查开盖区是否为空
        - 设置将球磨罐体放罐到开盖区
        - 等待放罐完成
        - 返回成功
        """
        logger.info(f"将研磨后球磨罐{mill_position}放到开盖区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: not (self.is_open_can_body_occupied())):
            error_msg = "开罐主体占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        pick_place_code = RoboticArmPickPlaceCode_1.OPEN_CAN_AFTER_MILL_PLACE_CAN_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开盖区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置开盖区放罐    
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}放到开盖区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}放到开盖区完成"): # 等待完成状态复位
                logger.info(f"将研磨后球磨罐{mill_position}放到开盖区完成")
                # 前端物料转移：将机械臂1暂存载具放入「开盖区」（无暂存则跳过）
                self._place_carrier_to_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": f"将研磨后球磨罐{mill_position}放到开盖区完成",
                }
            else:
                error_msg = f"将研磨后球磨罐{mill_position}放到开盖区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将研磨后球磨罐{mill_position}放到开盖区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_milled_can_from_open_can_position(self, mill_position: int) -> dict:
        """
        从开盖区抓取研磨后球磨罐
        - 检查机械臂1是否空闲
        - 检查开盖区是否为空
        - 设置将研磨后球磨罐从开盖区位置抓取
        - 等待抓取罐完成
        - 返回成功
        """
        logger.info(f"将研磨后球磨罐{mill_position}从开盖区位置抓取...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: self.is_open_can_body_occupied()):
            error_msg = "开罐主体未占位，无法抓取"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        pick_place_code = RoboticArmPickPlaceCode_1.OPEN_CAN_AFTER_MILL_PICK_BASE_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开盖区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置开盖区取座    
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}从开盖区位置抓取完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}从开盖区位置抓取完成"): # 等待完成状态复位
                logger.info(f"将研磨后球磨罐{mill_position}从开盖区位置抓取完成")
                # 前端物料转移：从「开盖区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": f"将研磨后球磨罐{mill_position}从开盖区位置抓取完成",
                }
            else:
                error_msg = f"将研磨后球磨罐{mill_position}从开盖区位置抓取失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将研磨后球磨罐{mill_position}从开盖区位置抓取失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_milled_can_to_sieve_position(self, mill_position: int) -> dict:
        """
        将研磨后球磨罐放到过筛区
        - 检查机械臂1是否空闲
        - 检查过筛区是否为空
        - 设置将研磨后球磨罐放到过筛区位置
        - 等待放到过筛区位置完成
        - 返回成功
        """
        logger.info(f"将研磨后球磨罐{mill_position}放到过筛区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: not (self.is_sieve_can_occupied())):
            error_msg = "过筛区球磨罐占位，无法放罐"
            logger.error(error_msg)
            raise ValueError(error_msg)

        pick_place_code = RoboticArmPickPlaceCode_1.SIEVE_PLACE_BASE_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.SIEVE_POSITION) # 设置机械臂目标位置为过筛区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置过筛区放座    
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}放到过筛区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}放到过筛区完成"): # 等待完成状态复位
                logger.info(f"将研磨后球磨罐{mill_position}放到过筛区完成")
                # 前端物料转移：将机械臂1暂存载具放入「过筛区」球磨罐位(1)（无暂存则跳过）
                self._place_carrier_to_warehouse_at("过筛区", "1", arm_id=1)
                return {
                    "success": True,
                    "message": f"将研磨后球磨罐{mill_position}放到过筛区完成",
                }
            else:
                error_msg = f"将研磨后球磨罐{mill_position}放到过筛区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将研磨后球磨罐{mill_position}放到过筛区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def sieve(self) -> dict:
        """
        过筛
        - 检查过筛占位
        - 设置过筛动作代码
        - 等待请求执行
        - 等待过筛完成
        - 返回成功
        """
        logger.info("过筛...")
        if not self._wait_condition(lambda: self.is_sieve_can_occupied()):
            error_msg = "过筛区球磨罐没有占位，无法过筛"
            logger.error(error_msg)
            raise ValueError(error_msg)

        if self._wait_until_true("Sieve_Request_Process", description="过筛请求"):
            logger.info("接收到过筛请求")
            self.set_node_value("Sieve_Action_Control_Code", SieveActionCode.SIEVE) # 设置过筛动作代码
            self.set_node_value("Sieve_Start_Process", True) # 设置过筛开始
            if self._wait_until_true("Sieve_Process_Complete", description="过筛完成"):
                logger.info("过筛完成")
                self.set_node_value("Sieve_Action_Control_Code", SieveActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Sieve_Start_Process", False) # 复位过筛开始
                return {
                    "success": True,
                    "message": "过筛完成",
                }
            else:
                logger.error("过筛失败，操作未完成")
                self.set_node_value("Sieve_Action_Control_Code", SieveActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Sieve_Start_Process", False) # 复位过筛开始
                raise ValueError("过筛失败，操作未完成")
        else:
            error_msg = "过筛失败，未收到过筛请求"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_milled_can_from_sieve_position(self, mill_position: int) -> dict:
        """
        从过筛区抓取研磨后球磨罐
        - 检查机械臂1是否空闲
        - 检查过筛区是否占位
        - 设置将研磨后球磨罐从过筛区位置抓取
        - 等待从过筛区位置抓取完成
        - 返回成功
        """
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: self.is_sieve_can_occupied()):
            error_msg = "过筛区球磨罐没有占位，无法从过筛区抓取球磨罐"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        pick_place_code = RoboticArmPickPlaceCode_1.SIEVE_PICK_BASE_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.SIEVE_POSITION) # 设置机械臂目标位置为过筛区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置过筛区取座       
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"从过筛区抓取研磨后球磨罐{mill_position}完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"从过筛区抓取研磨后球磨罐{mill_position}完成"): # 等待完成状态复位
                logger.info(f"从过筛区抓取研磨后球磨罐{mill_position}完成")
                # 前端物料转移：从「过筛区」球磨罐位(1)取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse_at("过筛区", "1", arm_id=1)
                return {
                    "success": True,
                    "message": f"从过筛区抓取研磨后球磨罐{mill_position}完成",
                }
            else:
                error_msg = f"从过筛区抓取研磨后球磨罐{mill_position}失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"从过筛区抓取研磨后球磨罐{mill_position}失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_milled_can_to_scrape_position(self, mill_position: int) -> dict:
        """
        将研磨后球磨罐放到刮粉区
        - 检查机械臂1是否空闲
        - 检查刮粉区是否占位
        - 设置将研磨后球磨罐放到刮粉区
        - 等待放到刮粉区完成
        - 返回成功
        """
        logger.info(f"将研磨后球磨罐{mill_position}放到刮粉区...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: not (self.is_scrape_occupied())):
            error_msg = "刮粉区占位，无法放罐"
            logger.error(error_msg)
            raise ValueError(error_msg)

        pick_place_code = RoboticArmPickPlaceCode_1.SCRAPE_POWDER_PLACE_BASE_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.SCRAPE_POWDER_POSITION) # 设置机械臂目标位置为刮粉区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置刮粉区放座    
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}放到刮粉区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}放到刮粉区完成"): # 等待完成状态复位
                logger.info(f"将研磨后球磨罐{mill_position}放到刮粉区完成")
                # 前端物料转移：将机械臂1暂存载具放入「刮粉区」（无暂存则跳过）
                self._place_carrier_to_warehouse("刮粉区", arm_id=1)
                return {
                    "success": True,
                    "message": f"将研磨后球磨罐{mill_position}放到刮粉区完成",
                }
            else:
                error_msg = f"将研磨后球磨罐{mill_position}放到刮粉失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将研磨后球磨罐{mill_position}放到刮粉区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def scrape_powder(self) -> dict:
        """
        刮粉
        - 检查刮粉区占位
        - 设置刮粉区动作代码
        - 等待请求执行
        - 等待刮粉区完成
        - 返回成功
        """
        logger.info("刮粉区...")
        if not self._wait_condition(lambda: self.is_scrape_occupied()):
            error_msg = "刮粉区没有占位，无法刮粉"
            logger.error(error_msg)
            raise ValueError(error_msg)

        if self._wait_until_true("Scrape_Powder_Request_Process", description="刮粉请求"):
            logger.info("接收到刮粉请求")
            self.set_node_value("Scrape_Powder_Action_Control_Code", ScrapePowderActionCode.SCRAPE_POWDER) # 设置刮粉区动作代码
            self.set_node_value("Scrape_Powder_Start_Process", True) # 设置刮粉开始
            if self._wait_until_true("Scrape_Powder_Process_Complete", description="刮粉完成"):
                logger.info("刮粉完成")
                self.set_node_value("Scrape_Powder_Action_Control_Code", ScrapePowderActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Scrape_Powder_Start_Process", False) # 复位刮粉开始
                return {
                    "success": True,
                    "message": "刮粉完成",
                }
            else:
                logger.error("刮粉失败，操作未完成")
                self.set_node_value("Scrape_Powder_Action_Control_Code", ScrapePowderActionCode.NO_ACTION) # 复位动作
                self.set_node_value("Scrape_Powder_Start_Process", False) # 复位刮粉开始
                raise ValueError("刮粉失败，操作未完成")
        else:
            error_msg = "刮粉失败，未收到刮粉请求"
            logger.error(error_msg)
            raise ValueError(error_msg)
        

    @action()
    def pick_milled_can_from_scrape_position(self, mill_position: int) -> dict:
        """
        从刮粉区取下研磨后球磨罐
        - 检查机械臂1是否空闲
        - 检查刮粉区是否占位
        - 设置将研磨后球磨罐从刮粉区取下
        - 等待从刮粉区取下完成
        - 返回成功
        """
        logger.info(f"从刮粉区位置取下研磨后球磨罐{mill_position}...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: self.is_scrape_occupied()):
            error_msg = "刮粉区没有占位，无法取下"
            logger.error(error_msg)
            raise ValueError(error_msg)

        pick_place_code = RoboticArmPickPlaceCode_1.SCRAPE_POWDER_PICK_BASE_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.SCRAPE_POWDER_POSITION) # 设置机械臂目标位置为刮粉区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置刮粉区取座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}从刮粉区取下完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将研磨后球磨罐{mill_position}从刮粉区取下完成"): # 等待完成状态复位
                logger.info(f"将研磨后球磨罐{mill_position}从刮粉区取下完成")
                # 前端物料转移：从「刮粉区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("刮粉区", arm_id=1)
                return {
                    "success": True,
                    "message": f"将研磨后球磨罐{mill_position}从刮粉区取下完成",
                }
            else:
                error_msg = f"将研磨后球磨罐{mill_position}从刮粉区取下失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将研磨后球磨罐{mill_position}从刮粉区取下失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def place_sieved_can_to_open_can_position(self, mill_position: int) -> dict:
        """
        将过筛后球磨罐放到开罐区位置
        - 检查机械臂1是否空闲
        - 检查过筛区是否占位
        - 设置将过筛后球磨罐放到开罐区位置
        - 等待放到开罐区位置完成
        - 返回成功
        """
        logger.info(f"将过筛后球磨罐{mill_position}放到开罐区位置...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: not (self.is_open_can_body_occupied())):
            error_msg = "开罐区占位，无法放罐"
            logger.error(error_msg)
            raise ValueError(error_msg)

        pick_place_code = RoboticArmPickPlaceCode_1.OPEN_CAN_AFTER_SIEVE_PLACE_BASE_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开罐区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置开罐区放座    
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将过筛后球磨罐{mill_position}放到开罐区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将过筛后球磨罐{mill_position}放到开罐区完成"): # 等待完成状态复位
                logger.info(f"将过筛后球磨罐{mill_position}放到开罐区完成")
                # 前端物料转移：将机械臂1暂存载具放入「开盖区」（无暂存则跳过）
                self._place_carrier_to_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": f"将过筛后球磨罐{mill_position}放到开罐区完成",
                }
            else:
                error_msg = f"将过筛后球磨罐{mill_position}放到开罐区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将过筛后球磨罐{mill_position}放到开罐区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def pick_sieved_can_from_open_can_position(self, mill_position: int) -> dict:
        """
        将过筛后球磨罐从开罐区位置取下
        - 检查机械臂1是否空闲
        - 检查开罐区是否占位
        - 设置将过筛后球磨罐从开罐区位置取下
        - 等待从开罐区位置取下完成
        - 返回成功
        """
        logger.info(f"将过筛后球磨罐{mill_position}从开罐区位置取下...")
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")
        
        if mill_position not in [1, 2, 3, 4]:
            error_msg = f"研磨后球磨罐编号{mill_position}无效"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: self.is_open_can_body_occupied()):
            error_msg = "开罐区没有占位，无法取下"
            logger.error(error_msg)
            raise ValueError(error_msg)

        pick_place_code = RoboticArmPickPlaceCode_1.OPEN_CAN_AFTER_SIEVE_PICK_CAN_1 + (mill_position - 1) * 10
        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.OPEN_CAN_POSITION) # 设置机械臂目标位置为开罐区
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", pick_place_code) # 设置开罐区取座
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将过筛后球磨罐{mill_position}从开罐区取下完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将过筛后球磨罐{mill_position}从开罐区取下完成"): # 等待完成状态复位
                logger.info(f"将过筛后球磨罐{mill_position}从开罐区取下完成")
                # 前端物料转移：从「开盖区」取走载具，暂存到机械臂1（无物料则跳过）
                self._pick_carrier_from_warehouse("开盖区", arm_id=1)
                return {
                    "success": True,
                    "message": f"将过筛后球磨罐{mill_position}从开罐区取下完成",
                }
            else:
                error_msg = f"将过筛后球磨罐{mill_position}从开罐区取下失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将过筛后球磨罐{mill_position}从开罐区取下失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def place_can_to_can_rack(self, rack_position: int) -> dict:
        """
        将球磨罐放到罐架区
        - 检查机械臂1是否空闲
        - 设置将球磨罐放到罐架位置
        - 等待将球磨罐放到罐架位置完成
        - 返回成功
        
        参数:
            rack_position: 罐架区位置
        
        Returns:
            dict: 包含 success 和 message
        """
        logger.info(f"放到罐架区放，位置：{rack_position}")
        MIN_RACK_POSITION = 1
        MAX_RACK_POSITION = RoboticArmPickPlaceCode_1.PLACE_CAN_RACK_END - RoboticArmPickPlaceCode_1.PLACE_CAN_RACK_START + 1
        if rack_position < MIN_RACK_POSITION or rack_position > MAX_RACK_POSITION:
            error_msg = "罐架位置错误"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self._wait_until_true("Robotic_Arm_Idle_1", description="等待机械臂1空闲")

        if not self._wait_condition(lambda: not (self.is_can_rack_occupied(rack_position))):
            error_msg = f"罐架位置{rack_position}已占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_1", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_1", RoboticArmTargetPosition_1.CAN_RACK_POSITION) # 设置机械臂目标位置为罐架
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_1", RoboticArmPickPlaceCode_1.PLACE_CAN_RACK_START + rack_position - 1) # 设置罐架位置
        self.set_node_value("Robotic_Arm_Action_Trigger_1", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_1", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_1", description=f"将球磨罐放到罐架位置{rack_position}完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_1", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_1", description=f"将球磨罐放到罐架位置{rack_position}完成"): # 等待完成状态复位
                logger.info(f"将球磨罐放到罐架位置{rack_position}完成")
                # 前端物料转移：将机械臂1暂存载具放入「球磨罐仓库」对应位（无暂存则跳过）
                self._place_carrier_to_warehouse_at("球磨罐仓库", self._can_rack_site_key(rack_position), arm_id=1)
                return {
                    "success": True,
                    "message": f"将球磨罐放到罐架位置{rack_position}完成",
                }
            else:
                error_msg = f"将球磨罐放到罐架位置{rack_position}失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将球磨罐放到罐架位置{rack_position}失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def pick_small_crucible_from_crucible_rack(self, rack_position: int) -> dict:
        """
        从坩锅架区取小坩埚
        - 检查机械臂2是否空闲
        - 设置取小坩埚位置
        - 等待取小坩埚位置完成
        - 返回成功
        """
        logger.info(f"从坩埚架区取小坩埚，位置：{rack_position}")
        MIN_RACK_POSITION = 1
        MAX_RACK_POSITION = RoboticArmPickPlaceCode_2.PICK_CRUCIBLE_RACK_END - RoboticArmPickPlaceCode_2.PICK_CRUCIBLE_RACK_START + 1
        if rack_position < MIN_RACK_POSITION or rack_position > MAX_RACK_POSITION:
            error_msg = "坩埚位置错误"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")
        
        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", RoboticArmPickPlaceCode_2.PICK_CRUCIBLE_RACK_START + rack_position - 1) # 设置坩埚位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"取小坩埚位置{rack_position}完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"取小坩埚位置{rack_position}完成"): # 等待完成状态复位
                logger.info(f"取小坩埚位置{rack_position}完成")
                # 前端物料转移：从「小坩埚仓库」对应位取走载具，暂存到机械臂2（无物料则跳过）
                self._pick_carrier_from_warehouse_at("小坩埚仓库", self._small_crucible_rack_site_key(rack_position), arm_id=2)
                return {
                    "success": True,
                    "message": f"取小坩埚位置{rack_position}完成",
                }
            else:
                error_msg = f"取小坩埚位置{rack_position}失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"取小坩埚位置{rack_position}失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def place_small_crucible_to_sieve_position(self) -> dict:
        """
        将小坩埚放到过筛区
        - 检查机械臂2是否空闲
        - 检查过筛是否占位
        - 设置放小坩埚位置
        - 等待放小坩锅位置完成
        - 返回成功
        """
        logger.info(f"将小坩埚放到过筛区")
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")
        
        if not self._wait_condition(lambda: not (self.is_sieve_crucible_occupied())):
            error_msg = "过筛区小坩锅已占位"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", RoboticArmPickPlaceCode_2.PLACE_SIEVE_CRUCIBLE) # 设置过筛区位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"放小坩埚到过筛区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"放小坩埚到过筛区完成"): # 等待完成状态复位
                logger.info(f"放小坩锅到过筛区完成")
                # 前端物料转移：将机械臂2暂存载具放入「过筛区」小坩埚位(3)（无暂存则跳过）
                self._place_carrier_to_warehouse_at("过筛区", "3", arm_id=2)
                return {
                    "success": True,
                    "message": f"放小坩埚到过筛区完成",
                }
            else:
                error_msg = f"放小坩埚到过筛区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"放小坩埚到过筛区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_funnel_from_crucible_rack(self, rack_position: int) -> dict:
        """
        从漏斗架区取漏斗
        - 检查机械臂2是否空闲
        - 设置取漏斗位置
        - 等待取漏斗完成
        - 返回成功
        """
        logger.info(f"从漏斗架区取漏斗，位置：{rack_position}")
        MIN_RACK_POSITION = 1
        MAX_RACK_POSITION = RoboticArmPickPlaceCode_2.PICK_FUNNEL_RACK_END - RoboticArmPickPlaceCode_2.PICK_FUNNEL_RACK_START + 1
        if rack_position < MIN_RACK_POSITION or rack_position > MAX_RACK_POSITION:
            error_msg = "漏斗架位置错误"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")

        funnel_pick_code = RoboticArmPickPlaceCode_2.PICK_FUNNEL_RACK_START + rack_position - 1
        if not self._wait_condition(lambda: self.is_crucible_rack_occupied(funnel_pick_code)):
            error_msg = f"漏斗架位置{rack_position}无漏斗，无法抓取"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", funnel_pick_code) # 设置漏斗架位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"取漏斗位置{rack_position}完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"取漏斗位置{rack_position}完成"): # 等待完成状态复位
                logger.info(f"取漏斗位置{rack_position}完成")
                # 前端物料转移：从「漏斗仓库」C-{rack_position} 取走载具，暂存到机械臂2（无物料则跳过）
                self._pick_carrier_from_warehouse_at("漏斗仓库", f"C-{rack_position}", arm_id=2)
                return {
                    "success": True,
                    "message": f"取漏斗位置{rack_position}完成",
                }
            else:
                error_msg = f"取漏斗位置{rack_position}失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"取漏斗位置{rack_position}失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_funnel_to_sieve_position(self) -> dict:
        """
        将漏斗放到过筛区
        - 检查机械臂2是否空闲
        - 检查过筛是否占位
        - 设置放漏斗位置
        - 等待放漏斗位置完成
        - 返回成功
        """
        logger.info(f"将漏斗放到过筛区")
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")
        
        if not self._wait_condition(lambda: not (self.is_sieve_funnel_occupied())):
            error_msg = "过筛区漏斗已占位"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", RoboticArmPickPlaceCode_2.PLACE_SIEVE_FUNNEL) # 设置过筛区位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"放漏斗到过筛区完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"放漏斗到过筛区完成"): # 等待完成状态复位
                logger.info(f"放漏斗到过筛区完成")
                # 前端物料转移：将机械臂2暂存载具放入「过筛区」漏斗位(2)（无暂存则跳过）
                self._place_carrier_to_warehouse_at("过筛区", "2", arm_id=2)
                return {
                    "success": True,
                    "message": f"放漏斗到过筛区完成",
                }
            else:
                error_msg = f"放漏斗到过筛区失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"放漏斗到过筛区失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_small_crucible_from_sieve_position(self) -> dict:
        """
        将小坩埚从过筛区取出
        - 检查机械臂2是否空闲
        - 检查过筛是否占位
        - 设置放小坩埚位置
        - 等待放小坩锅位置完成
        - 返回成功
        """
        logger.info(f"将小坩埚从过筛区取出")
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")
        
        if not self._wait_condition(lambda: self.is_sieve_crucible_occupied()):
            error_msg = "过筛区小坩埚未占位"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", RoboticArmPickPlaceCode_2.PICK_SIEVE_CRUCIBLE) # 设置过筛区位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"从过筛区取小坩埚完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"从过筛区取小坩埚完成"): # 等待完成状态复位
                logger.info(f"从过筛区取小坩埚完成")
                # 前端物料转移：从「过筛区」小坩埚位(3)取走载具，暂存到机械臂2（无物料则跳过）
                self._pick_carrier_from_warehouse_at("过筛区", "3", arm_id=2)
                return {
                    "success": True,
                    "message": f"从过筛区取小坩埚完成",
                }
            else:
                error_msg = f"从过筛区取小坩锅失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"从过筛区取小坩锅失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_small_crucible_to_moving_position(self, moving_position: int)  -> dict:
        """
        将小坩锅放到搬运位置
        - 检查机械臂2是否空闲
        - 检查移动位置是否在放料位
        - 设置放小坩埚位置
        - 等待放小坩埚位置完成
        - 返回成功
        """
        logger.info(f"将小坩埚放到搬运位置 {moving_position}")

        MIN_MOVING_POSITION = 1
        MAX_MOVING_POSITION = RoboticArmPickPlaceCode_2.PLACE_SMALL_CRUCIBLE_4 - RoboticArmPickPlaceCode_2.PLACE_SMALL_CRUCIBLE_1 + 1

        if not (MIN_MOVING_POSITION <= moving_position <= MAX_MOVING_POSITION):
            error_msg = f"搬运位置 {moving_position} 超出范围"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")
        
        if self.get_small_crucible_discharge_current_position() != SmallCrucibleDischargePosition.FEEDIFNG:
            error_msg = f"当前小坩锅搬运位置不是放料位，无法放到搬运位置"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", RoboticArmPickPlaceCode_2.PLACE_SMALL_CRUCIBLE_1 + moving_position - 1) # 设置搬运位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"将小坩锅放到搬运位置 {moving_position} 完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"将小坩锅放到搬运位置 {moving_position} 完成"): # 等待完成状态复位
                logger.info(f"将小坩锅放到搬运位置 {moving_position} 完成")
                # 前端物料转移：将机械臂2暂存载具放入「小坩埚出料」对应位（无暂存则跳过）
                self._place_carrier_to_warehouse_at("小坩埚出料", str(moving_position), arm_id=2)
                return {
                    "success": True,
                    "message": f"将小坩锅放到搬运位置 {moving_position} 完成",
                }
            else:
                error_msg = f"将小坩锅放到搬运位置 {moving_position} 失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"将小坩锅放到搬运位置 {moving_position} 失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_funnel_from_sieve_position(self) -> dict:
        """
        将漏斗从过筛区取出
        - 检查机械臂2是否空闲
        - 检查过筛是否占位
        - 设置放漏斗位置
        - 等待放漏斗位置完成
        - 返回成功
        """
        logger.info(f"将漏斗从过筛区取出")
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")
        
        if not self._wait_condition(lambda: self.is_sieve_funnel_occupied()):
            error_msg = "过筛区漏斗未占位"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", RoboticArmPickPlaceCode_2.PICK_SIEVE_FUNNEL) # 设置过筛区位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"从过筛区取漏斗完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"从过筛区取漏斗完成"): # 等待完成状态复位
                logger.info(f"从过筛区取漏斗完成")
                # 前端物料转移：从「过筛区」漏斗位(2)取走载具，暂存到机械臂2（无物料则跳过）
                self._pick_carrier_from_warehouse_at("过筛区", "2", arm_id=2)
                return {
                    "success": True,
                    "message": f"从过筛区取漏斗完成",
                }
            else:
                error_msg = f"从过筛区取漏斗失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"从过筛区取漏斗失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_funnel_to_crucible_rack(self, rack_position: int) -> dict:
        """
        将漏斗放到漏斗架
        - 检查机械臂2是否空闲
        - 设置取漏斗位置
        - 等待取漏斗完成
        - 返回成功
        """
        logger.info(f"将漏斗放到漏斗架，位置：{rack_position}")
        MIN_RACK_POSITION = 1
        MAX_RACK_POSITION = RoboticArmPickPlaceCode_2.PLACE_FUNNEL_RACK_END - RoboticArmPickPlaceCode_2.PLACE_FUNNEL_RACK_START + 1
        if rack_position < MIN_RACK_POSITION or rack_position > MAX_RACK_POSITION:
            error_msg = "漏斗架位置错误"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self._wait_until_true("Robotic_Arm_Idle_2", description="等待机械臂2空闲")

        funnel_place_code = RoboticArmPickPlaceCode_2.PLACE_FUNNEL_RACK_START + rack_position - 1
        if not self._wait_condition(lambda: not (self.is_crucible_rack_occupied(funnel_place_code))):
            error_msg = f"漏斗架位置{rack_position}已占位，无法放置"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_2", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_2", funnel_place_code) # 设置漏斗架位置
        self.set_node_value("Robotic_Arm_Action_Trigger_2", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_2", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_2", description=f"放漏斗位置{rack_position}完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_2", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_2", description=f"放漏斗位置{rack_position}完成"): # 等待完成状态复位
                logger.info(f"放漏斗位置{rack_position}完成")
                # 前端物料转移：将机械臂2暂存载具放入「漏斗仓库」D-{rack_position}（无暂存则跳过）
                self._place_carrier_to_warehouse_at("漏斗仓库", f"D-{rack_position}", arm_id=2)
                return {
                    "success": True,
                    "message": f"放漏斗位置{rack_position}完成",
                }
            else:
                error_msg = f"放漏斗位置{rack_position}失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"放漏斗位置{rack_position}失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        

    @action()
    def small_crucible_discharge(self) -> dict:
        """
        小坩锅出料
        - 检查 4 个出料占位都为 True
        - 设置出料操作
        - 等待出料完成
        - 返回成功
        """
        logger.info("小坩埚出料")

        if not self._wait_condition(lambda: all(self.is_small_crucible_discharge_occupied(i) for i in (1, 2, 3, 4))):
            unoccupied = [i for i in (1, 2, 3, 4) if not self.is_small_crucible_discharge_occupied(i)]
            error_msg = f"小坩埚出料占位 {unoccupied} 未占位，无法出料"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Small_Crucible_Discharge_Target_Position_Code", SmallCrucibleDischargePosition.DISCHARGE) # 设置出料位置
        self.set_node_value("Small_Crucible_Discharge_Action_Trigger", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Small_Crucible_Discharge_Action_Trigger", True) # 设置动作触发
        if self._wait_until_true("Small_Crucible_Discharge_Action_Complete", description=f"小坩埚出料完成"):
            self.set_node_value("Small_Crucible_Discharge_Action_Trigger", False) # 复位动作触发
            if self._wait_until_false("Small_Crucible_Discharge_Action_Complete", description=f"小坩埚出料完成"): # 等待完成状态复位
                logger.info(f"小坩埚出料完成")
                return {
                    "success": True,
                    "message": f"小坩埚出料完成",
                }
            else:
                error_msg = f"小坩埚出料失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"小坩埚出料失败，操作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def small_crucible_feed(self) -> dict:
        """
        小坩锅上料
        - 设置上料操作
        - 等待上料完成
        - 返回成功
        """
        logger.info("小坩埚上料")

        self.set_node_value("Small_Crucible_Discharge_Target_Position_Code", SmallCrucibleDischargePosition.FEEDING) # 设置上料位置
        self.set_node_value("Small_Crucible_Discharge_Action_Trigger", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Small_Crucible_Discharge_Action_Trigger", True) # 设置动作触发
        if self._wait_until_true("Small_Crucible_Discharge_Action_Complete", description=f"小坩埚上料完成"):
            self.set_node_value("Small_Crucible_Discharge_Action_Trigger", False) # 复位动作触发
            if self._wait_until_false("Small_Crucible_Discharge_Action_Complete", description=f"小坩埚上料完成"): # 等待完成状态复位
                logger.info(f"小坩埚上料完成，完成复位超时")
                return {
                    "success": True,
                    "message": f"小坩埚出料完成",
                }
            else:
                error_msg = f"小坩埚上料失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"小坩埚上料失败，操作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def large_crucible_discharge(self) -> dict:
        """
        大坩锅搬运位出料
        - 设置出料操作
        - 等待出料完成
        - 返回成功
        """
        logger.info("大坩埚出料")

        self.set_node_value("Large_Crucible_Feed_Target_Position_Code", LargeCrucibleFeedPosition.FEEDING) # 设置出料位置
        self.set_node_value("Large_Crucible_Feed_Action_Trigger", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Large_Crucible_Feed_Action_Trigger", True) # 设置动作触发
        if self._wait_until_true("Large_Crucible_Feed_Action_Complete", description=f"大坩埚出料完成"):
            self.set_node_value("Large_Crucible_Feed_Action_Trigger", False) # 复位动作触发
            if self._wait_until_false("Large_Crucible_Feed_Action_Complete", description=f"大坩锅出料完成"): # 等待完成状态复位
                logger.info(f"大坩锅出料完成")
                return {
                    "success": True,
                    "message": f"大坩埚出料完成",
                }
            else:
                error_msg = f"大坩埚出料失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"大坩锅出料失败，操作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def large_crucible_feed(self) -> dict:
        """
        大坩锅搬运位置上料
        - 设置上料操作
        - 等待上料完成
        - 返回成功
        """
        logger.info("大坩埚上料")

        self.set_node_value("Large_Crucible_Feed_Target_Position_Code", LargeCrucibleFeedPosition.PICKING) # 设置取料位置
        self.set_node_value("Large_Crucible_Feed_Action_Trigger", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Large_Crucible_Feed_Action_Trigger", True) # 设置动作触发
        if self._wait_until_true("Large_Crucible_Feed_Action_Complete", description=f"大坩锅上料完成"):
            self.set_node_value("Large_Crucible_Feed_Action_Trigger", False) # 复位动作触发
            if self._wait_until_false("Large_Crucible_Feed_Action_Complete", description=f"大坩锅上料完成"): # 等待完成状态复位
                logger.info(f"大坩锅上料完成")
                return {
                    "success": True,
                    "message": f"大坩锅上料完成",
                }
            else:
                error_msg = f"大坩锅上料失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"大坩锅上料失败，操作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_large_crucible_from_moving_position(self) -> dict:
        """
        从搬运区取大坩埚
        - 检查机械臂3是否空闲
        - 检查大坩埚入料是否在取料区
        - 设置取大坩埚
        - 等待取大坩锅完成
        - 返回成功
        """
        logger.info("从搬运区取大坩埚")
        self._wait_until_true("Robotic_Arm_Idle_3", description="等待机械臂3空闲")

        if self.get_large_crucible_feed_current_position() != LargeCrucibleFeedPosition.PICKING:
            error_msg = f"当前大坩埚搬运位置不是取料位，无法取料"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.set_node_value("Robotic_Arm_Action_Complete_3", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_3", RoboticArmTargetPosition_3.LARGE_CRUCIBLE_POSITION) # 设置机械臂目标位置为大坩埚
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_3", RoboticArmPickPlaceCode_3.PICK_FEED_LARGE_CRUCIBLE) # 设置取大坩埚
        self.set_node_value("Robotic_Arm_Action_Trigger_3", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_3", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_3", description=f"取大坩埚完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_3", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_3", description=f"取大坩锅完成"): # 等待完成状态复位
                logger.info(f"取大坩埚完成")
                # 前端物料转移：从「大坩埚入料」取走载具，暂存到机械臂3（无物料则跳过）
                self._pick_carrier_from_warehouse("大坩埚入料", arm_id=3)
                return {
                    "success": True,
                    "message": f"取大坩埚完成",
                }
            else:
                error_msg = f"取大坩埚失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"取大坩埚失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_large_crucible_to_muffle_furnace(self, muffle_furnace_position: int) -> dict:
        """
        把大坩埚放到马弗炉
        - 检查机械臂3是否空闲
        - 检查马弗炉是否占位
        - 设置放大坩埚
        - 等待放大坩锅完成
        - 返回成功
        """
        logger.info(f"把大坩锅放到马弗炉位置{muffle_furnace_position}")

        MIN_MUFFLE_FURNACE_POSITION = 1
        MAX_MUFFLE_FURNACE_POSITION = 6
        if muffle_furnace_position < MIN_MUFFLE_FURNACE_POSITION or muffle_furnace_position > MAX_MUFFLE_FURNACE_POSITION:
            error_msg = f"马弗炉位置{muffle_furnace_position}不在有效范围内"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self._wait_until_true("Robotic_Arm_Idle_3", description="等待机械臂3空闲")
        
        if not self._wait_condition(lambda: not (self.is_muffle_furnace_occupied(muffle_furnace_position))):
            error_msg = f"马弗炉位置{muffle_furnace_position}占位，无法放料"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_3", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_3", RoboticArmTargetPosition_3.MUFFLE_FURNACE_1_POSITION + muffle_furnace_position - 1) # 设置机械臂目标位置为马弗炉
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_3", RoboticArmPickPlaceCode_3.PLACE_MUFFLE_FURNACE_1 + muffle_furnace_position - 1) # 设置放马弗炉
        self.set_node_value("Robotic_Arm_Action_Trigger_3", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_3", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_3", description=f"放马弗炉完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_3", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_3", description=f"放马弗炉完成"): # 等待完成状态复位
                logger.info(f"放马弗炉完成")
                # 前端物料转移：将机械臂3暂存载具放入「马弗炉{muffle_furnace_position}」（无暂存则跳过）
                self._place_carrier_to_warehouse(f"马弗炉{muffle_furnace_position}", arm_id=3)
                return {
                    "success": True,
                    "message": f"放大坩埚到马弗炉{muffle_furnace_position}完成",
                }
            else:
                error_msg = f"放马弗炉失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"放马弗炉失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def muffle_furnace_sintering(self, muffle_furnace_position: int) -> dict:
        """
        马弗炉烧结
        - 检查马弗炉是否占位
        - 设置开始烧结
        - 等待烧结完成
        - 返回成功
        """
        logger.info(f"开始马弗炉{muffle_furnace_position}烧结")
        MIN_MUFFLE_FURNACE_POSITION = 1
        MAX_MUFFLE_FURNACE_POSITION = 6
        if muffle_furnace_position < MIN_MUFFLE_FURNACE_POSITION or muffle_furnace_position > MAX_MUFFLE_FURNACE_POSITION:
            error_msg = f"马弗炉位置{muffle_furnace_position}不在有效范围内"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if not self._wait_condition(lambda: self.is_muffle_furnace_occupied(muffle_furnace_position)):
            error_msg = f"马弗炉位置{muffle_furnace_position}未占位，无法烧结"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if self._wait_until_true(f"Muffle_Furnace_Request_Process_{muffle_furnace_position}", description=f"马弗炉{muffle_furnace_position}开始请求"):
            self.set_node_value(f"Muffle_Furnace_Start_Process_{muffle_furnace_position}", True) # 设置开始烧结
            if self._wait_until_true(f"Muffle_Furnace_Process_Complete_{muffle_furnace_position}", description=f"马弗炉{muffle_furnace_position}烧结完成"):
                logger.info(f"马弗炉{muffle_furnace_position}烧结完成")
                self.set_node_value(f"Muffle_Furnace_Start_Process_{muffle_furnace_position}", False) # 复位动作触发
                return {
                    "success": True,
                    "message": f"马弗炉{muffle_furnace_position}烧结完成",
                }
            else:
                logger.error(f"马弗炉{muffle_furnace_position}烧结失败")
                self.set_node_value(f"Muffle_Furnace_Start_Process_{muffle_furnace_position}", False) # 复位动作触发
                raise ValueError(f"马弗炉{muffle_furnace_position}烧结失败，等待烧结超时")
        else:
            error_msg = f"马弗炉{muffle_furnace_position}烧结失败，未收到请求"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def pick_large_crucible_from_muffle_furnace(self, muffle_furnace_position: int) -> dict:
        """
        从马弗炉取大坩埚
        - 检查机械臂3是否空闲
        - 检查马弗炉是否占位
        - 设置开始取
        - 等待取完成
        - 返回成功
        """
        logger.info(f"从马弗炉{muffle_furnace_position}取大坩埚")

        MIN_MUFFLE_FURNACE_POSITION = 1
        MAX_MUFFLE_FURNACE_POSITION = 6
        if muffle_furnace_position < MIN_MUFFLE_FURNACE_POSITION or muffle_furnace_position > MAX_MUFFLE_FURNACE_POSITION:
            error_msg = f"马弗炉位置{muffle_furnace_position}不在有效范围内"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self._wait_until_true("Robotic_Arm_Idle_3", description="等待机械臂3空闲")
        
        if not self._wait_condition(lambda: self.is_muffle_furnace_occupied(muffle_furnace_position)):
            error_msg = f"马弗炉位置{muffle_furnace_position}未占位，无法取大坩埚"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_3", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_3", RoboticArmTargetPosition_3.MUFFLE_FURNACE_1_POSITION + muffle_furnace_position - 1) # 设置机械臂目标位置为马弗炉
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_3", RoboticArmPickPlaceCode_3.PICK_MUFFLE_FURNACE_1 + muffle_furnace_position - 1) # 设置放马弗炉
        self.set_node_value("Robotic_Arm_Action_Trigger_3", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_3", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_3", description=f"从马弗炉取大坩埚完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_3", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_3", description=f"从马弗炉取大坩埚完成"): # 等待完成状态复位
                logger.info(f"从马弗炉取大坩埚完成")
                # 前端物料转移：从「马弗炉{muffle_furnace_position}」取走载具，暂存到机械臂3（无物料则跳过）
                self._pick_carrier_from_warehouse(f"马弗炉{muffle_furnace_position}", arm_id=3)
                return {
                    "success": True,
                    "message": f"从马弗炉{muffle_furnace_position}取大坩埚完成",
                }
            else:
                error_msg = f"从马弗炉取大坩埚失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"从马弗炉取大坩埚失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_large_crucible_to_upper_product_rack(self) -> dict:
        """
        放大坩埚到成品出料上位置
        - 检查机械臂3是否空闲
        - 检查成品出料上位置是否占位
        - 设置放成品出料上位置
        - 等待放成品出料上位置完成
        - 返回成功
        """
        logger.info(f"放大坩埚到成品出料上位置")
        self._wait_until_true("Robotic_Arm_Idle_3", description="等待机械臂3空闲")
        
        if not self._wait_condition(lambda: not (self.is_upper_product_rack_occupied())):
            error_msg = f"上成品架占位，无法放料"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_3", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_3", RoboticArmTargetPosition_3.DISCHARGE_POSITION) # 设置机械臂目标位置为成品出料架
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_3", RoboticArmPickPlaceCode_3.PLACE_DISCHARGE_UPPER) # 设置放成品出料上位置
        self.set_node_value("Robotic_Arm_Action_Trigger_3", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_3", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_3", description=f"放成品出料上位置完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_3", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_3", description=f"放成品出料上位置完成"): # 等待完成状态复位
                logger.info(f"放成品出料上位置完成")
                # 前端物料转移：将机械臂3暂存载具放入「大坩埚出料」上位(1)（无暂存则跳过）
                self._place_carrier_to_warehouse_at("大坩埚出料", "1", arm_id=3)
                return {
                    "success": True,
                    "message": f"放成品出料上位置完成",
                }
            else:
                error_msg = f"放成品出料上位置失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"放成品出料上位置失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)


    @action()
    def place_large_crucible_to_lower_product_rack(self) -> dict:
        """
        放大坩埚到成品出料下位置
        - 检查机械臂3是否空闲
        - 检查成品出料下位置是否占位
        - 设置放成品出料下位置
        - 等待放成品出料下位置完成
        - 返回成功
        """
        logger.info(f"放大坩埚到成品出料下位置")
        self._wait_until_true("Robotic_Arm_Idle_3", description="等待机械臂3空闲")
        
        if not self._wait_condition(lambda: not (self.is_lower_product_rack_occupied())):
            error_msg = f"下成品架占位，无法放料"
            logger.error(error_msg)
            raise ValueError(error_msg)

        self.set_node_value("Robotic_Arm_Action_Complete_3", False)  # 先复位完成标志，避免读到上一次动作的完成
        self.set_node_value("Robotic_Arm_Target_Position_Code_3", RoboticArmTargetPosition_3.DISCHARGE_POSITION) # 设置机械臂目标位置为成品出料架
        self.set_node_value("Robotic_Arm_Target_Pick_Place_Code_3", RoboticArmPickPlaceCode_3.PLACE_DISCHARGE_LOWER) # 设置放成品出料下位置
        self.set_node_value("Robotic_Arm_Action_Trigger_3", False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value("Robotic_Arm_Action_Trigger_3", True) # 设置动作触发
        if self._wait_until_true("Robotic_Arm_Action_Complete_3", description=f"放成品出料下位置完成"):
            self.set_node_value("Robotic_Arm_Action_Trigger_3", False) # 复位动作触发
            if self._wait_until_false("Robotic_Arm_Action_Complete_3", description=f"放成品出料下位置完成"): # 等待完成状态复位
                logger.info(f"放成品出料下位置完成")
                # 前端物料转移：将机械臂3暂存载具放入「大坩埚出料」下位(2)（无暂存则跳过）
                self._place_carrier_to_warehouse_at("大坩埚出料", "2", arm_id=3)
                return {
                    "success": True,
                    "message": f"放成品出料下位置完成",
                }
            else:
                error_msg = f"放成品出料下位置失败，完成复位超时"
                logger.error(error_msg)
                raise ValueError(error_msg)
        else:
            error_msg = f"放成品出料下位置失败，机械臂动作未完成"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    
    @action()
    def trigger_all_process(self) -> dict:
        logger.info(f"触发所有过程...")
        time.sleep(1)
        logger.info(f"------ 首先进行1-4号位4个罐的加粉加珠球磨流程...")
        for rack_pos in range(1, 4):
            # 从罐架区取球磨罐
            ret = self.pick_can_from_can_rack(rack_pos)
            if not ret["success"]:
                return ret
            
            # 将空罐放到开盖区
            ret = self.place_empty_can_to_open_can_position()
            if not ret["success"]:
                return ret
            
            # 打开罐上盖
            ret = self.open_can_lid()
            if not ret["success"]:
                return ret
            
            # 从开盖区抓取空罐
            ret = self.pick_empty_can_from_open_can_position()
            if not ret["success"]:
                return ret
            
            # 将罐体放置到加粉区
            ret = self.place_can_to_add_powder_position()
            if not ret["success"]:
                return ret
            
            # 加粉
            ret = self.add_powder()
            if not ret["success"]:
                return ret
            
            # 从加粉区取罐体
            ret = self.pick_can_from_add_powder_position()
            if not ret["success"]:
                return ret
            
            # 将罐体放置到加珠区
            ret = self.place_can_to_add_bead_position()
            if not ret["success"]:
                return ret
            
            # 加珠
            ret = self.add_bead()
            if not ret["success"]:
                return ret
            
            # 从加珠区取罐体
            ret = self.pick_can_from_add_bead_position()
            if not ret["success"]:
                return ret
            
            # 将带有粉珠的球磨罐放置到开盖区
            ret = self.place_can_with_powder_and_bead_to_open_can_position()
            if not ret["success"]:
                return ret
            
            # 关盖
            ret = self.close_can_lid()
            if not ret["success"]:
                return ret
            
            # 从开盖区抓取带有粉珠的球磨罐
            ret = self.pick_can_with_powder_and_bead_from_open_can_position()
            if not ret["success"]:
                return ret
            
            # 将罐体放置到球磨区
            ret = self.place_can_to_ball_mill(rack_pos)
            if not ret["success"]:
                return ret
            
        # 球磨
        ret = self.ball_mill()
        if not ret["success"]:
            return ret
            
        logger.info(f"------ 进行4个罐的过筛刮粉流程...")
        # 小坩埚搬运位上料
        self.small_crucible_feed()
        
        for mill_pos in range(1, 4):
            # 从球磨区取罐体
            ret = self.pick_can_from_ball_mill(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将研磨后球磨罐放到开盖区
            ret = self.place_milled_can_to_open_can_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 打开罐上盖
            ret = self.open_can_lid()
            if not ret["success"]:
                return ret
            
            # 从开盖区抓取研磨后球磨罐
            ret = self.pick_milled_can_from_open_can_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将研磨后球磨罐放到过筛区
            ret = self.place_milled_can_to_sieve_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 从坩锅架区取小坩埚
            self.pick_small_crucible_from_crucible_rack(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将小坩埚放到过筛区
            self.place_small_crucible_to_sieve_position()
            if not ret["success"]:
                return ret
            
            # 从漏斗架区取漏斗
            self.pick_funnel_from_crucible_rack(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将漏斗放到过筛区
            self.place_funnel_to_sieve_position()
            if not ret["success"]:
                return ret
            
            # 过筛
            ret = self.sieve()
            if not ret["success"]:
                return ret
            
            # 从过筛区抓取研磨后球磨罐
            self.pick_milled_can_from_sieve_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将研磨后球磨罐放到刮粉区
            ret = self.place_milled_can_to_scrape_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 刮粉
            ret = self.scrape_powder()
            if not ret["success"]:
                return ret
            
            # 从刮粉区位置取下研磨后球磨罐
            ret = self.pick_milled_can_from_scrape_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 再次放到过筛区
            ret = self.place_milled_can_to_sieve_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 再次过筛
            ret = self.sieve()
            if not ret["success"]:
                return ret
            
            # 从过筛区抓取研磨后球磨罐
            self.pick_milled_can_from_sieve_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将过筛后球磨罐放到开罐区位置
            self.place_sieved_can_to_open_can_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 关盖
            ret = self.close_can_lid()
            if not ret["success"]:
                return ret
            
            # 将过筛后球磨罐从开罐区位置取下
            ret = self.pick_sieved_can_from_open_can_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将球磨罐放到罐架区
            ret = self.place_can_to_can_rack(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将小坩埚从过筛区取出
            ret = self.pick_small_crucible_from_sieve_position()
            if not ret["success"]:
                return ret

            # 将小坩锅放到搬运位置
            ret = self.place_small_crucible_to_moving_position(mill_pos)
            if not ret["success"]:
                return ret
            
            # 将漏斗从过筛区取出
            ret = self.pick_funnel_from_sieve_position()
            if not ret["success"]:
                return ret
            
            # 将漏斗放到漏斗架
            ret = self.place_funnel_to_crucible_rack(mill_pos)
            if not ret["success"]:
                return ret
            
        # 小坩埚搬运位出料
        self.small_crucible_discharge()
        
        # 大坩埚出料
        self.large_crucible_discharge()
        
        # 人工操作，把小坩埚取下，并放置在大坩埚中
        logger.info("------ 人工操作，把小坩锅取下，并放置在大坩埚中...")
        time.sleep(30)

        logger.info("------ 进行马弗炉烧结工作...")
        # 大坩埚上料
        self.large_crucible_feed()

        muffle_furnace_pos = 1

        # 从搬运区取大坩埚
        ret = self.pick_large_crucible_from_moving_position()
        if not ret["success"]:
            return ret
        
        # 把大坩埚放到马弗炉
        ret = self.place_large_crucible_to_muffle_furnace(muffle_furnace_pos)
        if not ret["success"]:
            return ret
        
        # 马弗炉烧结
        ret = self.muffle_furnace_sintering(muffle_furnace_pos)
        if not ret["success"]:
            return ret
        
        # 从马弗炉取大坩埚
        ret = self.pick_large_crucible_from_muffle_furnace(muffle_furnace_pos)
        if not ret["success"]:
            return ret
        
        # 放大坩埚到成品出料上位置
        ret = self.place_large_crucible_to_upper_product_rack()
        if not ret["success"]:
            return ret
        
        logger.info("------ 成品已放置在出料位，整体流程结束")

        return {
            "success": True,
            "message": f"整体流程运行完成",
        } 

    @action()
    def set_muffle_furnace_params(self, param_file: str) -> dict:
        """
        设置马弗炉烧结参数（6 台分别设置）。

        从 Excel 参数文件读取并下发到各马弗炉写节点 马弗炉_写[N].<参数名>。
        Excel 含若干 sheet，每个 sheet 对应一台马弗炉（sheet 名中的数字 1~6 即炉号）；
        每个 sheet 两列：第一列"参数名"，第二列"参数值"，首行为表头；参数值为空的行会被跳过。
        参数名需与节点字段一致（见 templates/马弗炉参数模板.xlsx）。

        Args:
            param_file[马弗炉参数文件]: 马弗炉参数 Excel(.xlsx) 文件路径，含 6 个 sheet 分别设置 6 台马弗炉。
        """
        import re
        import openpyxl

        if param_file:
            param_file = param_file.strip().strip('"').strip("'")  # 去除可能的首尾引号/空白
        if not param_file or not os.path.isfile(param_file):
            error_msg = f"马弗炉参数文件不存在: {param_file}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        try:
            wb = openpyxl.load_workbook(param_file, data_only=True)
        except Exception as e:
            error_msg = f"无法打开马弗炉参数文件 {param_file}: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        total_written = 0
        per_furnace = {}
        errors = []
        for sheet in wb.worksheets:
            m = re.search(r"\d+", sheet.title)
            if not m:
                logger.warning(f"跳过无法识别炉号的 sheet: {sheet.title}")
                continue
            furnace_idx = int(m.group())
            if furnace_idx < 1 or furnace_idx > 6:
                logger.warning(f"跳过无效炉号 {furnace_idx} 的 sheet: {sheet.title}")
                continue

            written = 0
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:  # 跳过表头
                    continue
                if not row or row[0] is None or str(row[0]).strip() == "":
                    continue
                param_name = str(row[0]).strip()
                value = row[1] if len(row) > 1 else None
                if value is None or str(value).strip() == "":
                    continue  # 参数值为空则跳过该参数
                node_name = f"马弗炉_写[{furnace_idx}].{param_name}"
                try:
                    if self.set_node_value(node_name, int(float(value))):
                        written += 1
                    else:
                        errors.append(f"{node_name} 写入失败")
                except Exception as e:
                    errors.append(f"{node_name} 写入出错: {e}")

            per_furnace[furnace_idx] = written
            total_written += written

            # 写入了参数才触发该炉参数下发，等待下发完成并复位
            if written > 0:
                self._send_param_handshake(
                    f"Muffle_Furnace_Parameter_Send_{furnace_idx}",
                    f"Muffle_Furnace_Parameter_Send_Complete_{furnace_idx}",
                    description=f"马弗炉{furnace_idx}参数下发",
                )
            logger.info(f"马弗炉{furnace_idx} 参数下发完成，共 {written} 项")

        if total_written == 0:
            error_msg = f"马弗炉参数下发失败，未写入任何参数（文件: {param_file}）"
            logger.error(error_msg)
            raise ValueError(error_msg)

        return {
            "success": True,
            "message": f"马弗炉参数下发完成，共写入 {total_written} 项",
            "data": {"total_written": total_written, "per_furnace": per_furnace},
            "error": errors,
        }

    @action()
    def set_ball_mill_params(self, param_file: str) -> dict:
        """
        设置球磨工艺参数。

        从 Excel 参数文件读取并下发到球磨写节点 球磨工艺参数[1].<参数名>。
        球磨仅 1 台，读取第一个 sheet，两列：第一列"参数名"，第二列"参数值"，
        首行为表头；参数值为空的行会被跳过。
        参数名需与节点字段一致（见 templates/球磨参数模板.xlsx）。

        Args:
            param_file[球磨参数文件]: 球磨参数 Excel(.xlsx) 文件路径。
        """
        import openpyxl

        if param_file:
            param_file = param_file.strip().strip('"').strip("'")  # 去除可能的首尾引号/空白
        if not param_file or not os.path.isfile(param_file):
            error_msg = f"球磨参数文件不存在: {param_file}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        try:
            wb = openpyxl.load_workbook(param_file, data_only=True)
        except Exception as e:
            error_msg = f"无法打开球磨参数文件 {param_file}: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        sheet = wb.worksheets[0]  # 球磨仅 1 台，取第一个 sheet
        written = 0
        errors = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # 跳过表头
                continue
            if not row or row[0] is None or str(row[0]).strip() == "":
                continue
            param_name = str(row[0]).strip()
            value = row[1] if len(row) > 1 else None
            if value is None or str(value).strip() == "":
                continue  # 参数值为空则跳过该参数
            node_name = f"球磨工艺参数[1].{param_name}"
            try:
                if self.set_node_value(node_name, int(float(value))):
                    written += 1
                else:
                    errors.append(f"{node_name} 写入失败")
            except Exception as e:
                errors.append(f"{node_name} 写入出错: {e}")

        if written == 0:
            error_msg = f"球磨参数下发失败，未写入任何参数（文件: {param_file}）"
            logger.error(error_msg)
            raise ValueError(error_msg)

        # 触发球磨参数下发，等待下发完成并复位
        self._send_param_handshake(
            "Ball_Mill_Parameter_Send",
            "Ball_Mill_Parameter_Send_Complete",
            description="球磨参数下发",
        )

        logger.info(f"球磨参数下发完成，共 {written} 项")
        return {
            "success": True,
            "message": f"球磨参数下发完成，共写入 {written} 项",
            "data": {"written": written},
            "error": errors,
        }

    # ============ 加样参数：xlsx 结构定义（供解析 / 模板生成 / 档案生成共用） ============
    # 单值参数：sheet 名 → 支持的参数列表 [(节点名, 数据类型, caster, 说明)]
    # 合并"本罐参数" + "工艺单值"到单个"单值参数" sheet，共 6 项。
    _ADD_POWDER_SINGLE_SHEETS: dict = {
        "单值参数": [
            # —— 本罐参数（每次加样通常都会变） ——
            ("粉末名称",             "STRING", lambda s: str(s).strip(),  "本罐粉末名称"),
            ("加样_位置号",          "INT16",  lambda s: int(float(s)),   "加样位置号"),
            ("加样_重量",            "FLOAT",  lambda s: float(s),        "加样目标重量（克）"),
            # —— 工艺单值参数 ——
            ("加样_1ML震荡最高速度",  "INT16",  lambda s: int(float(s)),   "1ML 震荡最高速度"),
            ("加样_500NL震荡最高速度","INT16",  lambda s: int(float(s)),   "500NL 震荡最高速度"),
            ("加样_粉末号",           "INT16",  lambda s: int(float(s)),   "加样粉末号"),
        ],
    }
    # 数组参数：<基础名>[sheet_idx]（sheet 名为整数索引，默认给出 0~4 五组）
    _ADD_POWDER_ARRAY_BASES: list = [
        ("加样_1ML开口量",       "INT16", lambda s: int(float(s))),
        ("加样_1ML落粉均速",     "FLOAT", lambda s: float(s)),
        ("加样_1ML旋转速度",     "INT16", lambda s: int(float(s))),
        ("加样_1ML提前停止量",   "INT32", lambda s: int(float(s))),
        ("加样_500NL开口量",     "INT16", lambda s: int(float(s))),
        ("加样_500NL落粉均速",   "FLOAT", lambda s: float(s)),
        ("加样_500NL旋转速度",   "INT16", lambda s: int(float(s))),
        ("加样_500NL提前停止量", "INT32", lambda s: int(float(s))),
    ]
    _ADD_POWDER_ARRAY_INDICES: list = [0, 1, 2, 3, 4]

    @action()
    def set_add_powder_params(
        self,
        param_file: str = "",
        record_dir: str = "",
        check_can_occupied: bool = True,
    ) -> dict:
        """
        设置加样参数（一次性从 xlsx 下发所有参数）。

        本动作把加样所需的**全部**参数（单值参数 + 数组参数）
        统一从一份 Excel(.xlsx) 文件读取并下发到 PLC；不再接收任何参数入参。

        xlsx 结构（sheet 名严格如下）:
          - "单值参数"    ：3 列 [参数名 | 参数值 | 数据类型]，行支持 6 项：
                            粉末名称 / 加样_位置号 / 加样_重量
                            加样_1ML震荡最高速度 / 加样_500NL震荡最高速度 / 加样_粉末号
          - "0" ~ "4"     ：3 列 [参数名 | 参数值 | 数据类型]，每个 sheet 名 = 数组索引，
                            对应 8 项加样数组参数（1ML/500NL 各 4 个）：
                            加样_1ML开口量 / 加样_1ML落粉均速 / 加样_1ML旋转速度 / 加样_1ML提前停止量
                            加样_500NL开口量 / 加样_500NL落粉均速 / 加样_500NL旋转速度 / 加样_500NL提前停止量
                            写入节点为 <基础名>[<sheet 索引>]。
          - "_" 前缀 sheet（如 "_readme_"、"_meta_"）会被忽略；单元格为空的行会跳过；
            全部为空 → 不做任何写入、不触发握手。

        前置等待（与 add_powder 保持一致）：
          - xlsx 中存在任一非空参数时，在首次实际写入 OPC 之前，会先：
            1) （可选）等待罐体占位（is_add_sample_occupied；未占位则抛错）
            2) 等待 Add_Sample_Request_Process = True（无超时）
          - 全部参数为空 → 不做前置等待，直接返回。

        参数模板见 `templates/加样参数模板.xlsx`。档案 xlsx（保存在 record_dir 或默认
        records/ 目录）与输入模板结构同构，可直接作为下次调用的 param_file 回放。

        add_powder 只负责触发加粉动作，不再接收参数；请先调本动作把参数下发完成，再调 add_powder。

        Args:
            param_file[加样参数文件]: 加样参数 Excel(.xlsx) 文件绝对路径（留空 = 不下发任何参数）。
            record_dir[档案保存目录]: 本次下发档案 xlsx 的保存目录（可选; 留空 = 本模块下的 records/）。
            check_can_occupied[是否检查罐体占位]: True=下发前等待并校验罐体占位；False=跳过占位检查。
        """
        import re
        import openpyxl

        written = 0
        errors = []

        # 快照收集（用于生成档案 xlsx，也用于回放）
        xlsx_written = {}      # {sheet_idx: {base_name: coerced_value}}
        single_written = {}    # {sheet_title: {node_name: coerced_value}}
        # 解析出的原始入参（写档案时也保留一份，便于人工审阅）
        parsed_single_raw = {}  # {sheet_title: {node_name: raw_str}}

        # --- 解析 xlsx（若未提供，则视为空参数请求，直接返回） ---
        if not param_file or not str(param_file).strip():
            logger.info("未提供 param_file，跳过参数下发")
            return {
                "success": True,
                "message": "未提供 param_file，未做任何写入",
                "data": {"written": 0, "record_file": None},
                "error": errors,
            }
        param_file = param_file.strip().strip('"').strip("'")
        if not os.path.isfile(param_file):
            error_msg = f"加样参数文件不存在: {param_file}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        try:
            wb = openpyxl.load_workbook(param_file, data_only=True)
        except Exception as e:
            error_msg = f"无法打开加样参数文件 {param_file}: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        # 每个 sheet 按名字分派
        single_sheet_specs = self._ADD_POWDER_SINGLE_SHEETS
        array_bases = {name: (dtype, caster) for name, dtype, caster in self._ADD_POWDER_ARRAY_BASES}

        # ---- 预扫：xlsx 是否有任何可识别 sheet 的非空目标参数值 ----
        def _has_any_nonempty_target():
            for _sheet in wb.worksheets:
                _title = str(_sheet.title).strip()
                if not _title or _title.startswith("_"):
                    continue
                # 目标 sheet：单值 sheet 或数字 sheet
                if _title in single_sheet_specs:
                    pass
                else:
                    try:
                        int(_title)
                    except ValueError:
                        continue
                for _r_idx, _row in enumerate(_sheet.iter_rows(values_only=True), start=1):
                    if _r_idx == 1:
                        continue
                    if not _row or _row[0] is None or str(_row[0]).strip() == "":
                        continue
                    _v = _row[1] if len(_row) > 1 else None
                    if _v is not None and str(_v).strip() != "":
                        return True
            return False

        # ---- 前置等待：与 add_powder 一致（可选占位 + Add_Sample_Request_Process）----
        # 只在 xlsx 有实际要下发的参数时执行；全空 xlsx 直接跳过等待。
        if _has_any_nonempty_target():
            if check_can_occupied:
                logger.info("加样参数下发前置：等待罐体占位 + 加样请求加工...")
                if not self._wait_condition(lambda: self.is_add_sample_occupied()):
                    error_msg = "没有罐体，无法下发加样参数"
                    logger.error(error_msg)
                    raise ValueError(error_msg)
            else:
                logger.info("加样参数下发前置：跳过罐体占位检查，等待加样请求加工...")
            self._wait_until_true("Add_Sample_Request_Process", description="加样请求加工")
            logger.info("已收到加样请求加工，开始下发参数")
        else:
            logger.info("xlsx 中未检测到任何非空参数，跳过前置等待与参数下发")

        for sheet in wb.worksheets:
            title = str(sheet.title).strip()
            if not title or title.startswith("_"):
                continue  # _meta_ / _readme_ 等元信息 sheet 跳过

            # ---- 单值参数 sheet ----
            if title in single_sheet_specs:
                spec_by_name = {node: (dtype, caster, note) for node, dtype, caster, note in single_sheet_specs[title]}
                parsed_single_raw[title] = {}
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_idx == 1:
                        continue  # 表头
                    if not row or row[0] is None or str(row[0]).strip() == "":
                        continue
                    node_name = str(row[0]).strip()
                    value = row[1] if len(row) > 1 else None
                    if value is None or str(value).strip() == "":
                        continue
                    spec = spec_by_name.get(node_name)
                    if spec is None:
                        errors.append(f"未知单值参数名: {node_name} (sheet={title})")
                        continue
                    _dtype, caster, _note = spec
                    raw_s = str(value).strip()
                    parsed_single_raw[title][node_name] = raw_s
                    try:
                        coerced = caster(raw_s)
                        logger.info(
                            f"[加样参数下发] 写入 {node_name} = {coerced!r} "
                            f"(sheet={title} 原始值={value!r})"
                        )
                        if self.set_node_value(node_name, coerced):
                            written += 1
                            single_written.setdefault(title, {})[node_name] = coerced
                        else:
                            errors.append(f"{node_name} 写入失败")
                    except ValueError:
                        errors.append(f"{node_name} 无法解析: {value!r}")
                    except Exception as e:
                        errors.append(f"{node_name} 写入出错: {e}")
                continue

            # ---- 数组参数 sheet（sheet 名 = 索引） ----
            try:
                sheet_idx = int(title)
            except ValueError:
                logger.warning(f"跳过无法识别的 sheet: {title!r}（既不在单值 sheet 集合中，也不是整数索引）")
                continue

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    continue
                if not row or row[0] is None or str(row[0]).strip() == "":
                    continue
                raw_name = str(row[0]).strip()
                base_name = re.sub(r"\[\d+\]$", "", raw_name)  # 允许带 [N] 后缀
                node_name = f"{base_name}[{sheet_idx}]"
                value = row[1] if len(row) > 1 else None
                if value is None or str(value).strip() == "":
                    continue
                spec = array_bases.get(base_name)
                if spec is None:
                    errors.append(f"未知数组参数名: {raw_name} (sheet={title})")
                    continue
                _dtype, caster = spec
                try:
                    coerced = caster(str(value).strip())
                    logger.info(
                        f"[加样参数下发] 写入 {node_name} = {coerced!r} "
                        f"(xlsx原始值={value!r} type={type(value).__name__})"
                    )
                    if self.set_node_value(node_name, coerced):
                        written += 1
                        xlsx_written.setdefault(sheet_idx, {})[base_name] = coerced
                    else:
                        errors.append(f"{node_name} 写入失败")
                except Exception as e:
                    errors.append(f"{node_name} 写入出错: {e}")

        # 有任意参数写入 → 触发参数下发握手；全空 → 直接返回
        if written == 0:
            logger.info("加样参数全部为空，未做任何写入")
            return {
                "success": True,
                "message": "加样参数全部为空，未做任何写入",
                "data": {"written": 0, "record_file": None},
                "error": errors,
            }

        handshake_ok = self._send_param_handshake(
            "Add_Sample_Parameter_Send",
            "Add_Sample_Parameter_Send_Complete",
            description="加样参数下发",
        )
        logger.info(f"加样参数下发完成，共 {written} 项")

        # 存档：把本次下发的参数快照写入 <record_dir>/<时间>_<粉末名>.xlsx
        # record_dir 留空则用本模块下的 records/
        # 粉末名从解析出的 "单值参数" 中取；取不到时回退到 "no_powder"
        # 兼容旧的 "本罐参数" 命名（迁移期）
        powder_name_for_filename = ""
        try:
            for _title in ("单值参数", "本罐参数"):
                v = single_written.get(_title, {}).get("粉末名称", "") or \
                    parsed_single_raw.get(_title, {}).get("粉末名称", "")
                if v:
                    powder_name_for_filename = v
                    break
        except Exception:
            pass

        record_path = None
        try:
            record_path = self._dump_add_powder_snapshot(
                param_file=param_file,
                powder_name_for_filename=str(powder_name_for_filename or ""),
                parsed_single_raw=parsed_single_raw,
                xlsx_written=xlsx_written,
                single_written=single_written,
                written_count=written,
                errors=errors,
                handshake_ok=handshake_ok,
                record_dir=record_dir,
            )
            logger.info(f"加样参数存档已生成: {record_path}")
        except Exception as e:
            logger.warning(f"加样参数存档写入失败（不影响主流程）: {e}")

        return {
            "success": True,
            "message": f"加样参数下发完成，共写入 {written} 项",
            "data": {"written": written, "record_file": record_path},
            "error": errors,
        }

    def _dump_add_powder_snapshot(
        self,
        *,
        param_file: str,
        powder_name_for_filename: str,
        parsed_single_raw: dict,
        xlsx_written: dict,
        single_written: dict,
        written_count: int,
        errors: list,
        handshake_ok: bool,
        record_dir: str = "",
    ) -> str:
        """把本次加样参数下发的快照写入 <record_dir>/<时间>_<粉末名>.xlsx，返回文件绝对路径。

        - record_dir 为空 → 默认使用本模块下的 records/ 子目录；
        - record_dir 会做首尾空白和引号清洗；不存在则自动创建。

        档案 xlsx **结构与输入模板完全一致**（可直接作为下次 param_file 回放）：
        - "单值参数"、"0"~"4"：每个 sheet 3 列 [参数名, 参数值, 数据类型]
        - "_meta_"：时间/xlsx 源文件/写入项数/握手状态/错误列表（此 sheet 会被解析时忽略，不影响回放）
        """
        import re
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        now = datetime.now()
        ts = now.strftime("%Y%m%d_%H%M%S")

        # 文件名安全化：过滤 Windows 不允许的字符 \/:*?"<>|
        safe_powder = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", str(powder_name_for_filename or "").strip())
        if not safe_powder:
            safe_powder = "no_powder"
        # 防止过长
        safe_powder = safe_powder[:80]

        # 保存目录：入参 record_dir 优先；留空则用模块内 records/
        rd = str(record_dir or "").strip().strip('"').strip("'")
        if rd:
            records_dir = os.path.abspath(os.path.expanduser(rd))
        else:
            here = os.path.dirname(os.path.abspath(__file__))
            records_dir = os.path.join(here, "records")
        os.makedirs(records_dir, exist_ok=True)

        filename = f"{ts}_{safe_powder}.xlsx"
        dst = os.path.join(records_dir, filename)
        # 如果重名（同秒多次调用），追加序号
        counter = 1
        while os.path.exists(dst):
            filename = f"{ts}_{safe_powder}_{counter}.xlsx"
            dst = os.path.join(records_dir, filename)
            counter += 1

        # 样式
        header_fill = PatternFill("solid", fgColor="FF1A3A63")
        header_font = Font(bold=True, color="FFFFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        def _apply_header(ws, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # --- 单值参数 sheet（"单值参数"：6 项） ---
        # 已写入优先；否则用解析到的原始入参；否则留空
        for sheet_title, spec in self._ADD_POWDER_SINGLE_SHEETS.items():
            ws = wb.create_sheet(sheet_title)
            ws.append(["参数名", "参数值", "数据类型"])
            _apply_header(ws, 3)
            written_this = single_written.get(sheet_title, {})
            raw_this = parsed_single_raw.get(sheet_title, {})
            for node, dtype, _caster, _note in spec:
                if node in written_this:
                    val = written_this[node]
                elif node in raw_this:
                    val = raw_this[node]
                else:
                    val = ""
                ws.append([node, val, dtype])
            for i, w in enumerate([26, 18, 12], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

        # --- 数组参数 sheet（0~4 及所有出现过的索引） ---
        indices = sorted(set(list(xlsx_written.keys()) + list(self._ADD_POWDER_ARRAY_INDICES)))
        for idx in indices:
            ws = wb.create_sheet(str(idx))
            ws.append(["参数名", "参数值", "数据类型"])
            _apply_header(ws, 3)
            written_this_sheet = xlsx_written.get(idx, {})
            for name, dtype, _caster in self._ADD_POWDER_ARRAY_BASES:
                ws.append([name, written_this_sheet.get(name, ""), dtype])
            for i, w in enumerate([26, 18, 12], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

        # --- Sheet: _meta_（放最后；解析时被 _ 前缀过滤，不影响回放） ---
        ws = wb.create_sheet("_meta_")
        rows = [
            ("下发时间",       now.strftime("%Y-%m-%d %H:%M:%S")),
            ("xlsx 源文件",    param_file or "(未提供)"),
            ("写入项数",       written_count),
            ("参数下发握手",   "成功" if handshake_ok else "失败"),
            ("错误数量",       len(errors)),
        ]
        ws["A1"] = "字段"
        ws["B1"] = "值"
        _apply_header(ws, 2)
        for i, (k, v) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)
        if errors:
            ws.cell(row=len(rows) + 3, column=1, value="errors").font = Font(bold=True)
            for j, err in enumerate(errors, start=len(rows) + 4):
                ws.cell(row=j, column=1, value=err)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 60

        wb.save(dst)
        return dst

    def _send_param_handshake(self, send_node: str, complete_node: str, description: str) -> bool:
        """参数下发握手：上升沿触发下发 → 等待下发完成 → 复位触发并等待完成复位。

        与其它动作的触发/等待完成/复位写法保持一致。
        """
        self.set_node_value(send_node, False)  # 上升沿: 先复位
        time.sleep(0.5)
        self.set_node_value(send_node, True)  # 触发参数下发
        if self._wait_until_true(complete_node, description=f"{description}完成"):
            self.set_node_value(send_node, False)  # 复位下发触发
            self._wait_until_false(complete_node, description=f"{description}完成复位")  # 等待完成状态复位
            return True
        return False

            
    def _wait_until_true(
        self,
        node_name: str,
        interval: float = 0.2,
        description: str = None
    ) -> bool:
        """等待布尔节点变为 True（无超时，持续轮询直到满足）"""
        desc = description or node_name
        logger.info(f"等待 {desc} 变为 True...")

        while True:
            if self.get_node_value(node_name, use_cache=True):
                logger.info(f"✓ {desc} 已变为 True")
                return True
            time.sleep(interval)

    def _wait_until_false(
        self,
        node_name: str,
        interval: float = 0.2,
        description: str = None
    ) -> bool:
        """等待布尔节点变为 False（无超时，持续轮询直到满足）"""
        desc = description or node_name
        logger.info(f"等待 {desc} 变为 False...")

        while True:
            if not self.get_node_value(node_name, use_cache=True):
                logger.info(f"✓ {desc} 已变为 False")
                return True
            time.sleep(interval)

    def _wait_for_nodes(
        self,
        conditions: dict,  # {node_name: target_value, ...}
        interval: float = 0.2
    ) -> bool:
        """等待多个节点同时满足条件（无超时，持续轮询直到满足）"""
        while True:
            all_met = all(
                self.get_node_value(name, use_cache=True) == target
                for name, target in conditions.items()
            )
            if all_met:
                return True
            time.sleep(interval)


if __name__ == '__main__':
    # 调试用法
    xuseDevice = XUSEDevice(
        url="opc.tcp://192.168.1.10:4840",
        # url="opc.tcp://127.0.0.1:48010",
        csv_path=os.path.dirname(os.path.abspath(__file__)) + "/xuse_variables.csv"
    )

    # 启动心跳
    # xuseDevice.start_heart_beat()

    logger.setLevel(logging.INFO)

    time.sleep(3)

    # 初始化工作站
    # xuseDevice.init_workstation()
    
    # 显示命令行，让用户通过选择序号来完成相应的操作
    # 如果带有参数，则序号和各参数之间均由空格分隔
    while True:
        print("请选择操作：")
        print("0 初始化")
        print("0-1 加样单元初始化")
        print("1-1 从罐架区取球磨罐")
        print("1-2 将空罐放到开盖区")
        print("1-3 打开罐上盖")
        print("1-4 从开盖区抓取空罐")
        print("1-5 将罐体放置到加粉区")
        print("1-6 加粉")
        print("1-7 从加粉区取罐体")
        print("1-8 将罐体放置到加珠区")
        print("1-9 加珠")
        print("1-10 从加珠区取罐体")
        print("1-11 将带有粉珠的球磨罐放置到开盖区")
        print("1-12 关盖")
        print("1-13 从开盖区抓取带有粉珠的球磨罐")
        print("1-14 将罐体放置到球磨区")
        print("1-15 球磨")
        print("1-16 从球磨区抓取罐体")
        print("1-17 将研磨后球磨罐放到开盖区")
        print("1-18 从开盖区抓取研磨后球磨罐")
        print("1-19 将研磨后球磨罐放到过筛区")
        print("1-20 过筛")
        print("1-21 从过筛区抓取研磨后球磨罐")
        print("1-22 将研磨后球磨罐放到刮粉区")
        print("1-23 刮粉")
        print("1-24 从刮粉区位置取下研磨后球磨罐")
        print("1-25 将过筛后球磨罐放到开罐区位置")
        print("1-26 将过筛后球磨罐从开罐区位置取下")
        print("1-27 将球磨罐放到罐架区")
        print("2-1 从坩埚架区取小坩埚")
        print("2-2 将小坩埚放到过筛区")
        print("2-3 从漏斗架区取漏斗")
        print("2-4 将漏斗放到过筛区")
        print("2-5 将小坩埚从过筛区取出")
        print("2-6 将小坩埚放到搬运位置")
        print("2-7 将漏斗从过筛区取出")
        print("2-8 将漏斗放到漏斗架")
        print("2-9 小坩埚搬运位出料")
        print("2-10 小坩埚搬运位上料")
        print("3-1 大坩埚搬运位出料")
        print("3-2 大坩埚搬运位置上料")
        print("3-3 从搬运区取大坩埚")
        print("3-4 把大坩埚放到马弗炉")
        print("3-5 马弗炉烧结")
        print("3-6 从马弗炉取大坩埚")
        print("3-7 放大坩埚到成品出料上位置")
        print("3-8 放大坩埚到成品出料下位置")
        print("98 全部流程")
        print("99 退出")
        choice = input("请输入操作序号：")
        if choice == "0":
            xuseDevice.trigger_init()
        elif choice == "0-1":
            xuseDevice.trigger_add_powder_init()
        elif choice.startswith("1-1 "):
            rack_pos = int(choice.split(" ")[1])
            xuseDevice.pick_can_from_can_rack(rack_pos)
        elif choice.startswith("1-2 "):
            xuseDevice.place_empty_can_to_open_can_position()
        elif choice.startswith("1-3 "):
            xuseDevice.open_can_lid()
        elif choice.startswith("1-4 "):
            xuseDevice.pick_empty_can_from_open_can_position()
        elif choice.startswith("1-5 "):
            xuseDevice.place_can_to_add_powder_position()
        elif choice.startswith("1-6 "):
            xuseDevice.add_powder()
        elif choice.startswith("1-7 "):
            xuseDevice.pick_can_from_add_powder_position()
        elif choice.startswith("1-8 "):
            xuseDevice.place_can_to_add_bead_position()
        elif choice.startswith("1-9 "):
            xuseDevice.add_bead()
        elif choice.startswith("1-10 "):
            xuseDevice.pick_can_from_add_bead_position()
        elif choice.startswith("1-11 "):
            xuseDevice.place_can_with_powder_and_bead_to_open_can_position()
        elif choice.startswith("1-12 "):
            xuseDevice.close_can_lid()
        elif choice.startswith("1-13 "):
            xuseDevice.pick_can_with_powder_and_bead_from_open_can_position()
        elif choice.startswith("1-14 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.place_can_to_ball_mill(mill_pos)
        elif choice.startswith("1-15 "):
            xuseDevice.ball_mill()
        elif choice.startswith("1-16 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.pick_can_from_ball_mill(mill_pos)
        elif choice.startswith("1-17 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.place_milled_can_to_open_can_position(mill_pos)
        elif choice.startswith("1-18 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.pick_milled_can_from_open_can_position(mill_pos)
        elif choice.startswith("1-19 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.place_milled_can_to_sieve_position(mill_pos)
        elif choice.startswith("1-20 "):
            xuseDevice.sieve()
        elif choice.startswith("1-21 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.pick_milled_can_from_sieve_position(mill_pos)
        elif choice.startswith("1-22 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.place_milled_can_to_scrape_position(mill_pos)
        elif choice.startswith("1-23 "):
            xuseDevice.scrape_powder()
        elif choice.startswith("1-24 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.pick_milled_can_from_scrape_position(mill_pos)
        elif choice.startswith("1-25 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.place_sieved_can_to_open_can_position(mill_pos)
        elif choice.startswith("1-26 "):
            mill_pos = int(choice.split(" ")[1])
            xuseDevice.pick_sieved_can_from_open_can_position(mill_pos)
        elif choice.startswith("1-27 "):
            rack_pos = int(choice.split(" ")[1])
            xuseDevice.place_can_to_can_rack(rack_pos)
        

        elif choice.startswith("2-1 "):
            rack_pos = int(choice.split(" ")[1])
            xuseDevice.pick_small_crucible_from_crucible_rack(rack_pos)
        elif choice.startswith("2-2 "):
            xuseDevice.place_small_crucible_to_sieve_position()
        elif choice.startswith("2-3 "):
            rack_pos = int(choice.split(" ")[1])
            xuseDevice.pick_funnel_from_crucible_rack(rack_pos)
        elif choice.startswith("2-4 "):
            xuseDevice.place_funnel_to_sieve_position()
        elif choice.startswith("2-5 "):
            xuseDevice.pick_small_crucible_from_sieve_position()
        elif choice.startswith("2-6 "):
            moving_pos = int(choice.split(" ")[1])
            xuseDevice.place_small_crucible_to_moving_position(moving_pos)
        elif choice.startswith("2-7 "):
            xuseDevice.pick_funnel_from_sieve_position()
        elif choice.startswith("2-8 "):
            rack_pos = int(choice.split(" ")[1])
            xuseDevice.place_funnel_to_crucible_rack(rack_pos)
        elif choice.startswith("2-9 "):
            xuseDevice.small_crucible_discharge()
        elif choice.startswith("2-10 "):
            xuseDevice.small_crucible_feed()


        elif choice.startswith("3-1 "):
            xuseDevice.large_crucible_discharge()
        elif choice.startswith("3-2 "):
            xuseDevice.large_crucible_feed()
        elif choice.startswith("3-3 "):
            xuseDevice.pick_large_crucible_from_moving_position()
        elif choice.startswith("3-4 "):
            furnace_pos = int(choice.split(" ")[1])
            xuseDevice.place_large_crucible_to_muffle_furnace(furnace_pos)
        elif choice.startswith("3-5 "):
            xuseDevice.muffle_furnace_sintering()
        elif choice.startswith("3-6 "):
            furnace_pos = int(choice.split(" ")[1])
            xuseDevice.pick_large_crucible_from_muffle_furnace(furnace_pos)
        elif choice.startswith("3-7 "):
            xuseDevice.place_large_crucible_to_upper_product_rack()
        elif choice.startswith("3-8 "):
            xuseDevice.place_large_crucible_to_lower_product_rack()


        elif choice.startswith("98 "):
            xuseDevice.trigger_all_process()
        elif choice.startswith("99 "):
            break
        else:
            print("无效的操作序号，请重新输入。")

    # 结束心跳
    # xuseDevice.stop_heart_beat()

    # 断开连接
    xuseDevice.disconnect()

    print("退出程序。")
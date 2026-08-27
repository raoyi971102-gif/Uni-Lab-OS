import csv
import importlib
from datetime import datetime
from pathlib import Path

import openpyxl

from unilabos.registry.decorators import get_topic_config


xuse_module = importlib.import_module("unilabos.devices.workstation.XUSE.XUSE")
XUSEDevice = xuse_module.XUSEDevice
XUSE_DIR = Path(xuse_module.__file__).resolve().parent


class CompletedPowderDevice(XUSEDevice):
    def __init__(self, *, powder_name, target_weight, actual_weight):
        self.values = {
            "Powder_Name": powder_name,
            "Add_Sample_Weight": target_weight,
            "Powder_Weight": actual_weight,
        }
        self.writes = []
        self.reads = []

    def _wait_until_true(self, *args, **kwargs):
        return True

    def _wait_until_false(self, *args, **kwargs):
        return True

    def set_node_value(self, node_name, value):
        self.writes.append((node_name, value))
        return True

    def get_node_value(self, node_name, **kwargs):
        self.reads.append((node_name, kwargs))
        return self.values[node_name]


def _csv_row(path: Path, name: str) -> dict:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        rows = [row for row in csv.DictReader(file) if row["Name"] == name]
    assert len(rows) == 1
    return rows[0]


def test_powder_weight_variable_is_in_real_and_sim_tables():
    real_row = _csv_row(XUSE_DIR / "xuse_variables.csv", "加粉重量")
    sim_row = _csv_row(XUSE_DIR / "xuse_variables_sim.csv", "加粉重量")

    assert real_row == {
        "Name": "加粉重量",
        "EnglishName": "Powder_Weight",
        "NodeType": "VARIABLE",
        "DataType": "FLOAT",
        "NodeLanguage": "Chinese",
        "NodeId": "ns=4;s=uniab|加粉重量",
    }
    assert sim_row["EnglishName"] == "Powder_Weight"
    assert sim_row["DataType"] == "FLOAT"
    assert sim_row["NodeId"] == "ns=2;s=XUSE.设备 1.加粉重量"


def test_powder_weight_is_uploaded_as_one_second_monitor():
    config = get_topic_config(XUSEDevice.powder_weight)
    device = object.__new__(XUSEDevice)
    device._powder_weight_cache = 1.2345

    assert config["period"] == 1.0
    assert config["name"] == "加粉重量"
    assert device.powder_weight() == 1.2345


def test_powder_weight_poller_force_reads_plc_each_second():
    class OnePassStop:
        stopped = False

        def is_set(self):
            return self.stopped

        def wait(self, seconds):
            assert seconds == 1.0
            self.stopped = True

    device = object.__new__(XUSEDevice)
    device._arm_status_nodes = []
    device._arm_status_poller_stop = OnePassStop()
    calls = []

    def get_node_value(node_name, **kwargs):
        calls.append((node_name, kwargs))
        return 3.1415

    device.get_node_value = get_node_value
    device._arm_status_poll_loop()

    assert device._powder_weight_cache == 3.1415
    assert calls == [
        ("Powder_Weight", {"use_cache": False, "force_read": True})
    ]


def test_add_powder_appends_actual_measurement_log(tmp_path):
    device = CompletedPowderDevice(
        powder_name="测试粉末_A",
        target_weight=1.25,
        actual_weight=1.2478,
    )

    first_result = device.add_powder(
        check_can_occupied=False,
        actual_powder_log_dir=f'"{tmp_path}"',
        can_number=12,
    )
    device.values.update(
        Powder_Name="测试粉末_B",
        Add_Sample_Weight=2.5,
        Powder_Weight=2.4931,
    )
    second_result = device.add_powder(
        check_can_occupied=False,
        actual_powder_log_dir=str(tmp_path),
        can_number=20,
    )

    log_path = tmp_path / "实际加粉日志.xlsx"
    assert first_result["data"]["actual_powder_log_file"] == str(log_path)
    assert second_result["data"]["actual_powder_log_file"] == str(log_path)
    assert first_result["can_number"] == 12
    assert second_result["data"]["can_number"] == 20
    assert device._powder_weight_cache == 2.4931
    assert all(kwargs == {"use_cache": False, "force_read": True} for _, kwargs in device.reads)

    workbook = openpyxl.load_workbook(log_path, data_only=True)
    sheet = workbook["实际加粉日志"]
    assert [cell.value for cell in sheet[1]] == [
        "时间戳",
        "加粉名称",
        "目标加粉重量",
        "实际加粉重量",
        "球磨罐编号",
    ]
    assert sheet.max_row == 3
    assert isinstance(sheet["A2"].value, datetime)
    assert [
        sheet["B2"].value,
        sheet["C2"].value,
        sheet["D2"].value,
        sheet["E2"].value,
    ] == [
        "测试粉末_A",
        1.25,
        1.2478,
        12,
    ]
    assert [
        sheet["B3"].value,
        sheet["C3"].value,
        sheet["D3"].value,
        sheet["E3"].value,
    ] == [
        "测试粉末_B",
        2.5,
        2.4931,
        20,
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:E3"
    workbook.close()


def test_add_powder_upgrades_legacy_four_column_log(tmp_path):
    log_path = tmp_path / "实际加粉日志.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "实际加粉日志"
    sheet.append(["时间戳", "加粉名称", "目标加粉重量", "实际加粉重量"])
    sheet.append([datetime(2026, 8, 27, 12, 0, 0), "历史粉末", 1.0, 0.99])
    workbook.save(log_path)
    workbook.close()

    device = CompletedPowderDevice(
        powder_name="新粉末",
        target_weight=2.0,
        actual_weight=1.98,
    )
    device.add_powder(
        check_can_occupied=False,
        actual_powder_log_dir=str(tmp_path),
        can_number=7,
    )

    workbook = openpyxl.load_workbook(log_path, data_only=True)
    sheet = workbook["实际加粉日志"]
    assert [cell.value for cell in sheet[1]] == [
        "时间戳",
        "加粉名称",
        "目标加粉重量",
        "实际加粉重量",
        "球磨罐编号",
    ]
    assert [sheet["B2"].value, sheet["C2"].value, sheet["D2"].value] == [
        "历史粉末",
        1.0,
        0.99,
    ]
    assert sheet["E2"].value is None
    assert [sheet["B3"].value, sheet["E3"].value] == ["新粉末", 7]
    assert sheet.auto_filter.ref == "A1:E3"
    workbook.close()
